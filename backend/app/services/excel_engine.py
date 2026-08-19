import io
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.styles import Font, Alignment

from app.core.config import settings
from app.models.report import Report
from app.models.photo import Photo

# Type-specific cell mappings per sheet matching docs/EXCEL_CELL_MAPPING.md and tools/generate_excel.py
TYPE_CELL_MAPPINGS: Dict[str, Dict[str, Dict[str, str]]] = {
    "HERMETIK": {
        "KAPAK SAYFASI": {
            "D9": "customer_name",
            "D10": "trafo_label",
            "D11": "address",
            "D12": "report_date",
            "D14": "test_date",
            "A29": "summary_text",
            "D55": "operator_title",
        },
        "ANA SAYFA": {
            "G11": "brand",
            "O11": "tap_info_1",
            "Q11": "tap_info_2",
            "S11": "tap_info_3",
            "G13": "power_kva",
            "O13": "manufacture_year",
            "G15": "voltage",
            "O15": "serial_no",
            "G17": "oil_brand",
            "G31": "dc_redresor_voltage",
            "O17": "oil_weight",
            "G19": "connection_group",
            "O19": "short_circuit_imp_pct",
            "G21": "tank_type",
            "I21": "tank_mark_hermetik",
            "P21": "tank_mark_gt",
            "U21": "tank_mark_kuru",
            "C55": "og_rab",
            "C57": "og_rbc",
            "C59": "og_rca",
            "J55": "ag_ran",
            "J57": "ag_rbn",
            "J59": "ag_rcn",
            "O55": "ag_rab",
            "O57": "ag_rbc",
            "O59": "ag_rca",
            "F66": "iso_yg_tank_30s",
            "I66": "iso_yg_tank_45s",
            "L66": "iso_yg_tank_60s",
            "F68": "iso_ag_tank_30s",
            "I68": "iso_ag_tank_45s",
            "L68": "iso_ag_tank_60s",
            "F70": "iso_yg_ag_30s",
            "I70": "iso_yg_ag_45s",
            "L70": "iso_yg_ag_60s",
            "B73": "notes",
            "F81": "operator_title",
        },
        "OG SARGI MEVCUT KADEME": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "AG SARGI": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "İZOLASYON ": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "D16": "iso_og_gnd",
            "D17": "iso_ag_gnd",
            "D30": "iso_temp",
            "D31": "iso_humidity",
        },
        "Ç.O 34500": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "B16": "ttr_tap1_a",
            "C16": "ttr_tap1_b",
            "D16": "ttr_tap1_c",
            "B17": "ttr_tap2_a",
            "C17": "ttr_tap2_b",
            "D17": "ttr_tap2_c",
            "B18": "ttr_tap3_a",
            "C18": "ttr_tap3_b",
            "D18": "ttr_tap3_c",
            "B19": "ttr_tap4_a",
            "C19": "ttr_tap4_b",
            "D19": "ttr_tap4_c",
            "B20": "ttr_tap5_a",
            "C20": "ttr_tap5_b",
            "D20": "ttr_tap5_c",
        },
        "TOPRAKLAMALAR": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D17": "ground_r_trafo_body",
            "D18": "ground_r_neutral",
            "D19": "ground_r_tank",
            "D32": "ground_r_og_lightning",
            "D33": "ground_r_panel",
            "D34": "ground_r_fence",
        },
        "HV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_hv_humidity",
        },
        "LV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_lv_humidity",
        },
        "ANA SAYFA KESİCİ": {
            "G11": "breaker_brand",
            "O11": "breaker_serial_no",
            "G13": "breaker_model",
            "O13": "breaker_year",
            "G15": "breaker_voltage",
            "G17": "breaker_motor_voltage",
            "O17": "breaker_coil_voltage",
            "G26": "breaker_dc_redresor_voltage",
            "G50": "breaker_iso_r_gnd",
            "O50": "breaker_contact_r",
            "D68": "breaker_notes",
            "F76": "operator_title",
        },
        "KESİCİ İZOLASYON": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "KESİCİ KONTAK": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "AÇMA-KAPAMA": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D10": "breaker_timing_open",
        },
        "DİĞER": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
        },
        "AKIM TRAFOLARI": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D16": "ct_ratio",
        },
        "HERMETİK YAĞ DİLEKÇESİ": {}
    },
    "KURU_TIP": {
        "KAPAK SAYFASI": {
            "D9": "customer_name",
            "D10": "trafo_label",
            "D11": "address",
            "D12": "report_date",
            "D14": "test_date",
            "A29": "summary_text",
            "D55": "operator_title",
        },
        "ANA SAYFA": {
            "G11": "brand",
            "O11": "tap_info_1",
            "Q11": "tap_info_2",
            "S11": "tap_info_3",
            "G13": "power_kva",
            "O13": "manufacture_year",
            "G15": "voltage",
            "O15": "serial_no",
            "G17": "connection_group",
            "G31": "dc_redresor_voltage",
            "O17": "short_circuit_imp_pct",
            "G19": "tank_type",
            "I19": "tank_mark_hermetik",
            "P19": "tank_mark_gt",
            "U19": "tank_mark_kuru",
            "C49": "og_rab",
            "C51": "og_rbc",
            "C53": "og_rca",
            "J49": "ag_ran",
            "J51": "ag_rbn",
            "J53": "ag_rcn",
            "O49": "ag_rab",
            "O51": "ag_rbc",
            "O53": "ag_rca",
            "C55": "ground_r_trafo_body",
            "F55": "ground_r_neutral",
            "J55": "ground_r_tank",
            "C57": "ground_r_og_lightning",
            "F57": "ground_r_panel",
            "J57": "ground_r_fence",
            "B67": "notes",
            "F74": "operator_title",
        },
        "OG SARGI MEVCUT KADEME": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "AG SARGI": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "İZOLASYON ": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "D16": "iso_og_gnd",
            "D17": "iso_ag_gnd",
            "D30": "iso_temp",
            "D31": "iso_humidity",
        },
        "Ç.O 34500": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "B16": "ttr_tap1_a",
            "C16": "ttr_tap1_b",
            "D16": "ttr_tap1_c",
            "B17": "ttr_tap2_a",
            "C17": "ttr_tap2_b",
            "D17": "ttr_tap2_c",
            "B18": "ttr_tap3_a",
            "C18": "ttr_tap3_b",
            "D18": "ttr_tap3_c",
            "B19": "ttr_tap4_a",
            "C19": "ttr_tap4_b",
            "D19": "ttr_tap4_c",
            "B20": "ttr_tap5_a",
            "C20": "ttr_tap5_b",
            "D20": "ttr_tap5_c",
        },
        "TOPRAKLAMALAR": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D17": "ground_r_trafo_body",
            "D18": "ground_r_neutral",
            "D19": "ground_r_tank",
            "D32": "ground_r_og_lightning",
            "D33": "ground_r_panel",
            "D34": "ground_r_fence",
        },
        "HV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_hv_humidity",
        },
        "LV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_lv_humidity",
        },
        "ANA SAYFA KESİCİ": {
            "G11": "breaker_brand",
            "O11": "breaker_serial_no",
            "G13": "breaker_model",
            "O13": "breaker_year",
            "G15": "breaker_voltage",
            "G17": "breaker_motor_voltage",
            "O17": "breaker_coil_voltage",
            "G26": "breaker_dc_redresor_voltage",
            "G50": "breaker_iso_r_gnd",
            "O50": "breaker_contact_r",
            "D68": "breaker_notes",
            "F76": "operator_title",
        },
        "KESİCİ İZOLASYON": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "KESİCİ KONTAK": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "AÇMA-KAPAMA": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D10": "breaker_timing_open",
        },
        "DİĞER": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
        },
        "AKIM TRAFOLARI": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D16": "ct_ratio",
        }
    },
    "GT": {
        "KAPAK SAYFASI": {
            "D9": "customer_name",
            "D10": "trafo_label",
            "D11": "address",
            "D12": "report_date",
            "D14": "test_date",
            "A29": "summary_text",
            "D55": "operator_title",
        },
        "ANA SAYFA": {
            "G11": "brand",
            "O11": "tap_info_1",
            "Q11": "tap_info_2",
            "S11": "tap_info_3",
            "G13": "power_kva",
            "O13": "manufacture_year",
            "G15": "voltage",
            "O15": "serial_no",
            "G17": "oil_brand",
            "G31": "dc_redresor_voltage",
            "O17": "oil_weight",
            "G19": "connection_group",
            "O19": "short_circuit_imp_pct",
            "G21": "tank_type",
            "I21": "tank_mark_hermetik",
            "P21": "tank_mark_gt",
            "U21": "tank_mark_kuru",
            "C55": "og_rab",
            "C57": "og_rbc",
            "C59": "og_rca",
            "J55": "ag_ran",
            "J57": "ag_rbn",
            "J59": "ag_rcn",
            "O55": "ag_rab",
            "O57": "ag_rbc",
            "O59": "ag_rca",
            "F66": "iso_yg_tank_30s",
            "I66": "iso_yg_tank_45s",
            "L66": "iso_yg_tank_60s",
            "F68": "iso_ag_tank_30s",
            "I68": "iso_ag_tank_45s",
            "L68": "iso_ag_tank_60s",
            "F70": "iso_yg_ag_30s",
            "I70": "iso_yg_ag_45s",
            "L70": "iso_yg_ag_60s",
            "J63": "ground_r_fence",
            "B73": "notes",
            "F81": "operator_title",
        },
        "OG SARGI MEVCUT KADEME": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "AG SARGI": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "İZOLASYON ": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "D16": "iso_og_gnd",
            "D17": "iso_ag_gnd",
            "D30": "iso_temp",
            "D31": "iso_humidity",
        },
        "Ç.O 34500": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "B16": "ttr_tap1_a",
            "C16": "ttr_tap1_b",
            "D16": "ttr_tap1_c",
            "B17": "ttr_tap2_a",
            "C17": "ttr_tap2_b",
            "D17": "ttr_tap2_c",
            "B18": "ttr_tap3_a",
            "C18": "ttr_tap3_b",
            "D18": "ttr_tap3_c",
            "B19": "ttr_tap4_a",
            "C19": "ttr_tap4_b",
            "D19": "ttr_tap4_c",
            "B20": "ttr_tap5_a",
            "C20": "ttr_tap5_b",
            "D20": "ttr_tap5_c",
        },
        "TOPRAKLAMALAR": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D17": "ground_r_trafo_body",
            "D18": "ground_r_neutral",
            "D19": "ground_r_tank",
            "D32": "ground_r_og_lightning",
            "D33": "ground_r_panel",
            "D34": "ground_r_fence",
        },
        "ANA SAYFA KESİCİ": {
            "G11": "breaker_brand",
            "O11": "breaker_serial_no",
            "G13": "breaker_model",
            "O13": "breaker_year",
            "G15": "breaker_voltage",
            "G17": "breaker_motor_voltage",
            "O17": "breaker_coil_voltage",
            "G26": "breaker_dc_redresor_voltage",
            "G50": "breaker_iso_r_gnd",
            "O50": "breaker_contact_r",
            "D68": "breaker_notes",
            "F76": "operator_title",
        },
        "KESİCİ İZOLASYON": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "KESİCİ KONTAK": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "AÇMA-KAPAMA": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D10": "breaker_timing_open",
        },
        "DİĞER": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
        },
        "AKIM TRAFOLARI": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D16": "ct_ratio",
        },
        "YAĞ RAPORU": {
            "D16": "oil_test_breakdown_voltage",
            "D18": "oil_test_water_content",
        },
        "HERMETİK YAĞ DİLEKÇESİ": {}
    }
}

SAMPLE_CLEAR_RANGES = {
    "ANA SAYFA": [
        "G31", "O55", "O57", "O59",
        "F66", "I66", "L66",
        "F68", "I68", "L68",
        "F70", "I70", "L70"
    ],
    "ANA SAYFA KESİCİ": [
        "G11", "O11", "G13", "O13", "G15", "O15", "G17", "O17",
        "G24", "I24", "P24", "R24", "G26", "H26", "I26", "P26", "R26",
        "G28", "I28", "P28", "R28", "G30", "I30", "P30", "R30",
        "G32", "I32", "P32", "R32",
        "C39", "F39", "I39", "L39", "O39",
        "C41", "F41", "I41", "L41", "O41",
        "C43", "F43", "I43", "L43", "O43",
        "C50", "G50", "K50", "O50",
        "C52", "G52", "K52", "O52",
        "C54", "G54", "K54", "O54",
        "C61", "J61", "Q61", "C63", "J63", "Q63",
        "C65", "F65", "J65", "M65", "Q65",
        "F77", "F78",
    ],
    "OG SARGI MEVCUT KADEME": [
        "D11", "J11", "O11",
        "C24", "E24", "G24", "I24", "K24", "M24", "O24", "Q24",
        "C25", "E25", "G25", "I25", "K25", "M25", "O25", "Q25",
        "C26", "E26", "G26", "I26", "K26", "M26", "O26", "Q26",
    ],
    "AG SARGI": [
        "D11", "J11", "O11",
        "C24", "E24", "G24", "I24", "K24", "M24", "O24", "Q24",
        "C25", "E25", "G25", "I25", "K25", "M25", "O25", "Q25",
        "C26", "E26", "G26", "I26", "K26", "M26", "O26", "Q26",
    ],
    "İZOLASYON ": [
        "D11", "J11", "O11", "D16", "D17", "D30", "D31",
    ],
    "Ç.O 34500": [
        "D11", "J11", "O11",
        "B16", "C16", "D16", "B17", "C17", "D17",
        "B18", "C18", "D18", "B19", "C19", "D19", "B20", "C20", "D20",
    ],
    "TOPRAKLAMALAR": [
        "D9", "J9", "O9", "D17", "D18", "D19", "D32", "D33", "D34",
    ],
    "HV PF": ["D11", "J11", "O11", "P17"],
    "LV PF": ["D11", "J11", "O11", "P17"],
    "KESİCİ İZOLASYON": ["D10", "J10", "O10"],
    "KESİCİ KONTAK": ["D10", "J10", "O10"],
    "AÇMA-KAPAMA": ["D9", "J9", "O9", "D10"],
    "DİĞER": ["D9", "J9", "O9", "D16", "H16", "D17", "H17", "D18", "H18", "D23", "H23", "D24", "H24", "D25", "H25", "D30", "H30", "D31", "H31", "D32", "H32"],
    "AKIM TRAFOLARI": ["D9", "J9", "O9", "D16", "H16", "D17", "H17", "D18", "H18"],
}

# Sheet signature anchors disabled across all sheets (Rule C)
SHEET_SIGNATURE_ANCHORS = {
    "HERMETIK": {},
    "KURU_TIP": {},
    "GT": {}
}

CHECKLIST_PAIRS = {
    "HERMETIK": {
        "checklist_1": {"evet": "I27", "hayir": "J27"},
        "checklist_2": {"evet": "I29", "hayir": "J29"},
        "checklist_3": {"evet": "I31", "hayir": "J31"},
        "checklist_4": {"evet": "I33", "hayir": "J33"},
        "checklist_5": {"evet": "I35", "hayir": "J35"},
        "checklist_6": {"evet": "I37", "hayir": "J37"},
        "checklist_7": {"evet": "I39", "hayir": "J39"},
        "checklist_8": {"evet": "I41", "hayir": "J41"},
        "checklist_9": {"evet": "R27", "hayir": "S27"},
        "checklist_10": {"evet": "R29", "hayir": "S29"},
        "checklist_11": {"evet": "R31", "hayir": "S31"},
        "checklist_12": {"evet": "R33", "hayir": "S33"},
        "checklist_13": {"evet": "R35", "hayir": "S35"},
        "checklist_14": {"evet": "R37", "hayir": "S37"},
        "checklist_15": {"evet": "R39", "hayir": "S39"},
        "checklist_16": {"evet": "R41", "hayir": "S41"},
        "breaker_control_visual": {"evet": "G24", "hayir": "I24"},
        "breaker_control_cleanliness": {"evet": "P24", "hayir": "R24"},
        "breaker_control_dc_redresor": {"evet": "I26", "hayir": "J26"},
        "breaker_control_cell_cleanliness": {"evet": "P26", "hayir": "R26"},
        "breaker_control_indicator": {"evet": "G28", "hayir": "I28"},
        "breaker_control_busbar": {"evet": "P28", "hayir": "R28"},
        "breaker_control_mechanical": {"evet": "G30", "hayir": "I30"},
        "breaker_control_heater": {"evet": "P30", "hayir": "R30"},
        "breaker_control_cable": {"evet": "G32", "hayir": "I32"},
        "breaker_control_relay": {"evet": "P32", "hayir": "R32"},
    },
    "KURU_TIP": {
        "checklist_1": {"evet": "I27", "hayir": "J27"},
        "checklist_2": {"evet": "I29", "hayir": "J29"},
        "checklist_3": {"evet": "I31", "hayir": "J31"},
        "checklist_4": {"evet": "I33", "hayir": "J33"},
        "checklist_5": {"evet": "I35", "hayir": "J35"},
        "checklist_6": {"evet": "I37", "hayir": "J37"},
        "checklist_7": {"evet": "R27", "hayir": "S27"},
        "checklist_8": {"evet": "R29", "hayir": "S29"},
        "checklist_9": {"evet": "R31", "hayir": "S31"},
        "checklist_10": {"evet": "R33", "hayir": "S33"},
        "checklist_11": {"evet": "R35", "hayir": "S35"},
        "checklist_12": {"evet": "R37", "hayir": "S37"},
        "breaker_control_visual": {"evet": "G24", "hayir": "I24"},
        "breaker_control_cleanliness": {"evet": "P24", "hayir": "R24"},
        "breaker_control_dc_redresor": {"evet": "I26", "hayir": "J26"},
        "breaker_control_cell_cleanliness": {"evet": "P26", "hayir": "R26"},
        "breaker_control_indicator": {"evet": "G28", "hayir": "I28"},
        "breaker_control_busbar": {"evet": "P28", "hayir": "R28"},
        "breaker_control_mechanical": {"evet": "G30", "hayir": "I30"},
        "breaker_control_heater": {"evet": "P30", "hayir": "R30"},
        "breaker_control_cable": {"evet": "G32", "hayir": "I32"},
        "breaker_control_relay": {"evet": "P32", "hayir": "R32"},
    },
    "GT": {
        "checklist_1": {"evet": "I27", "hayir": "J27"},
        "checklist_2": {"evet": "I29", "hayir": "J29"},
        "checklist_3": {"evet": "I31", "hayir": "J31"},
        "checklist_4": {"evet": "I33", "hayir": "J33"},
        "checklist_5": {"evet": "I35", "hayir": "J35"},
        "checklist_6": {"evet": "I37", "hayir": "J37"},
        "checklist_7": {"evet": "I39", "hayir": "J39"},
        "checklist_8": {"evet": "I41", "hayir": "J41"},
        "checklist_9": {"evet": "R27", "hayir": "S27"},
        "checklist_10": {"evet": "R29", "hayir": "S29"},
        "checklist_11": {"evet": "R31", "hayir": "S31"},
        "checklist_12": {"evet": "R33", "hayir": "S33"},
        "checklist_13": {"evet": "R35", "hayir": "S35"},
        "checklist_14": {"evet": "R37", "hayir": "S37"},
        "checklist_15": {"evet": "R39", "hayir": "S39"},
        "checklist_16": {"evet": "R41", "hayir": "S41"},
        "breaker_control_visual": {"evet": "G24", "hayir": "I24"},
        "breaker_control_cleanliness": {"evet": "P24", "hayir": "R24"},
        "breaker_control_dc_redresor": {"evet": "I26", "hayir": "J26"},
        "breaker_control_cell_cleanliness": {"evet": "P26", "hayir": "R26"},
        "breaker_control_indicator": {"evet": "G28", "hayir": "I28"},
        "breaker_control_busbar": {"evet": "P28", "hayir": "R28"},
        "breaker_control_mechanical": {"evet": "G30", "hayir": "I30"},
        "breaker_control_heater": {"evet": "P30", "hayir": "R30"},
        "breaker_control_cable": {"evet": "G32", "hayir": "I32"},
        "breaker_control_relay": {"evet": "P32", "hayir": "R32"},
    }
}


def pre_clear_sample_cells(wb):
    """
    Pre-clears hardcoded sample data cells across all sheets in wb.
    Preserves all formulas, titles, labels, logos, formatting, and row heights.
    """
    for sname, cell_list in SAMPLE_CLEAR_RANGES.items():
        matched_sheet = None
        for name in wb.sheetnames:
            if name.strip() == sname.strip():
                matched_sheet = name
                break
        if matched_sheet:
            ws = wb[matched_sheet]
            for cref in cell_list:
                try:
                    cell = ws[cref]
                    if type(cell).__name__ != 'MergedCell' and cell.value is not None:
                        if not str(cell.value).strip().startswith('='):
                            cell.value = None
                except Exception:
                    pass


def date_to_excel_serial(date_input: Any) -> Optional[float]:
    """Converts DateTime or string (DD.MM.YYYY or YYYY-MM-DD) into Excel serial date number."""
    if date_input is None:
        return None
    try:
        if isinstance(date_input, (int, float)):
            return float(date_input)

        s_date = str(date_input).strip()
        if not s_date:
            return None

        dt_obj = None
        if "." in s_date:
            parts = s_date.split(".")
            if len(parts) == 3:
                p1, p2, p3 = int(parts[0]), int(parts[1]), int(parts[2])
                if p3 > 1000:
                    dt_obj = datetime(p3, p2, p1)
                elif p1 > 1000:
                    dt_obj = datetime(p1, p2, p3)
        elif "-" in s_date:
            parts = s_date.split("-")
            if len(parts) == 3:
                dt_obj = datetime(int(parts[0]), int(parts[1]), int(parts[2]))

        if dt_obj is not None:
            epoch = datetime(1899, 12, 30)
            return float((dt_obj - epoch).days)
    except Exception:
        pass
    return None


def get_writable_cell(ws, cell_ref: str):
    """Returns the top-left cell if cell_ref belongs to a merged range."""
    cell = ws[cell_ref]
    if type(cell).__name__ == 'MergedCell':
        for rng in ws.merged_cells.ranges:
            if cell_ref in rng:
                return ws.cell(rng.min_row, rng.min_col)
    return cell


def process_checklist_pairs(ws, report_type, data_dict):
    """
    Sets Evet/Hayır checkmarks for checklist items and clears residual template sample checkmarks.
    Also clears 3.0 pt divider row cells C61, F61, J61, C63, F63, J63 on ANA SAYFA to prevent ghost text.
    """
    for cref in ["C61", "F61", "J61", "C63", "F63", "J63"]:
        try:
            cell = get_writable_cell(ws, cref)
            cell.value = None
        except Exception:
            pass

    pairs_map = CHECKLIST_PAIRS.get(report_type, CHECKLIST_PAIRS["HERMETIK"])
    for key, pair in pairs_map.items():
        evet_cell = get_writable_cell(ws, pair["evet"])
        hayir_cell = get_writable_cell(ws, pair["hayir"])

        val = data_dict.get(key)
        if val is None:
            evet_cell.value = None
            hayir_cell.value = None
        else:
            is_true = (val is True or str(val).strip().lower() in ["true", "ü", "1", "evet"])
            is_false = (val is False or str(val).strip().lower() in ["false", "0", "hayir", "hayır"])
            if is_true:
                evet_cell.value = "ü"
                hayir_cell.value = None
            elif is_false:
                evet_cell.value = None
                hayir_cell.value = "ü"
            else:
                evet_cell.value = None
                hayir_cell.value = None


def process_sheet_signature(ws, target_anchor: Optional[str], sig_path: Optional[str]):
    """
    Cleans sample text artifacts from the worksheet.
    DOES NOT insert any user signature image (signature placement is disabled across all sheets).
    Clears all right block cells below 'ONAYLAYAN' header.
    """
    for r in range(40, min(86, ws.max_row + 1)):
        for c in range(5, 15):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell' and cell.value is not None:
                cval_str = str(cell.value).strip()
                cval_u = cval_str.upper()
                if "ONAYLAYAN" in cval_u:
                    for br in range(r + 1, min(r + 7, ws.max_row + 1)):
                        for bc in range(c, min(c + 7, 20)):
                            bcell = ws.cell(row=br, column=bc)
                            if type(bcell).__name__ != 'MergedCell':
                                bcell.value = None


def process_kapak_photos(ws_kapak, data_dict, photos: Optional[List[Photo]] = None):
    """
    Inserts loaded photos (photo_before, photo_after, photo_label, photo_extra)
    into KAPAK SAYFASI and writes bold Turkish captions beneath each photo slot.
    """
    p_before = data_dict.get("photo_before")
    p_after = data_dict.get("photo_after")
    p_label = data_dict.get("photo_label")
    p_extra = data_dict.get("photo_extra")

    if photos:
        for p in photos:
            if hasattr(p, 'photo_type') and hasattr(p, 'file_path'):
                ptype = str(p.photo_type).lower()
                if 'before' in ptype:
                    p_before = p.file_path
                elif 'after' in ptype:
                    p_after = p.file_path
                elif 'label' in ptype or 'tag' in ptype:
                    p_label = p.file_path
                elif 'extra' in ptype:
                    p_extra = p.file_path

    photos_to_place = []
    if p_before and isinstance(p_before, str) and os.path.exists(p_before):
        photos_to_place.append((p_before, "Bakım Öncesi"))
    if p_after and isinstance(p_after, str) and os.path.exists(p_after):
        photos_to_place.append((p_after, "Bakım Sonrası"))
    if p_label and isinstance(p_label, str) and os.path.exists(p_label):
        photos_to_place.append((p_label, "Trafo Etiket / Plaka"))
    if p_extra and isinstance(p_extra, str) and os.path.exists(p_extra):
        photos_to_place.append((p_extra, "Ek Fotoğraf"))

    slots = [
        ("A32", "A38", 200, 120),
        ("G32", "G38", 200, 120),
        ("A40", "A47", 200, 115),
        ("G40", "G47", 200, 115),
    ]

    for cref in ["A38", "G38", "A41", "G41", "A47", "G47", "A50", "G50"]:
        try:
            cell = get_writable_cell(ws_kapak, cref)
            cell.value = None
        except Exception:
            pass

    font_bold = Font(name="Calibri", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    caption_merges = [("A38", "A38:D38"), ("G38", "G38:J38"), ("A47", "A47:D47"), ("G47", "G47:J47")]
    for cref, mrange in caption_merges:
        try:
            ws_kapak.merge_cells(mrange)
        except Exception:
            pass

    for idx, (p_path, label_text) in enumerate(photos_to_place):
        if idx < len(slots):
            anchor_cell, caption_ref, w_px, h_px = slots[idx]
            try:
                img = OpenPyXLImage(p_path)
                img.width = w_px
                img.height = h_px
                ws_kapak.add_image(img, anchor_cell)

                cell = get_writable_cell(ws_kapak, caption_ref)
                cell.value = label_text
                cell.font = font_bold
                cell.alignment = align_center
            except Exception:
                pass


def clean_5070_text(wb):
    """
    Clears any cell value containing '5070' or 'elektronik imza kanunu' across all sheets.
    """
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    cval_u = cell.value.upper()
                    if "5070" in cval_u or "ELEKTRONİK İMZA KANUNU" in cval_u or "ELEKTRONIK IMZA KANUNU" in cval_u:
                        cell.value = None


def clean_left_personnel_block(wb, op_name: str, op_title: str, test_date: Optional[str] = None, report_date: Optional[str] = None, data_dict: Optional[Dict[str, Any]] = None):
    """
    Ensures KAPAK sol kutu has ONLY:
      - Test Tarihi : {test_date}
      - Rapor Tarihi : {report_date}
      - Unvan : {operator_title}

    Cleans and standardizes personnel blocks and footers across ALL sub-pages:
      - UNVAN: written verbatim from operator_title
      - TEST TARİHİ: test_date
      - ODA SİCİL NO / EKİPNET NO: written ONLY if valid sicil_no / ekipnet_no exists in data_dict, otherwise CLEARED.
    """
    if data_dict is None:
        data_dict = {}

    real_sicil = str(data_dict.get("sicil_no") or "").strip()
    real_ekipnet = str(data_dict.get("ekipnet_no") or "").strip()

    if real_sicil == op_name or real_sicil == op_title:
        real_sicil = ""
    if real_ekipnet == op_name or real_ekipnet == op_title:
        real_ekipnet = ""

    for sname in wb.sheetnames:
        ws = wb[sname]
        if sname == "KAPAK SAYFASI":
            for c_let in ["A", "B", "C", "D", "E", "F"]:
                try:
                    get_writable_cell(ws, f"{c_let}52").value = None
                except Exception:
                    pass

            get_writable_cell(ws, "A53").value = "Test Tarihi"
            get_writable_cell(ws, "C53").value = ":"
            if test_date:
                s_val = date_to_excel_serial(test_date)
                cell_d53 = get_writable_cell(ws, "D53")
                cell_d53.value = s_val if s_val is not None else str(test_date)
                cell_d53.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "A54").value = "Rapor Tarihi"
            get_writable_cell(ws, "C54").value = ":"
            if report_date:
                s_val = date_to_excel_serial(report_date)
                cell_d54 = get_writable_cell(ws, "D54")
                cell_d54.value = s_val if s_val is not None else str(report_date)
                cell_d54.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "A55").value = "Unvan"
            get_writable_cell(ws, "C55").value = ":"
            get_writable_cell(ws, "D55").value = op_title

            for r in [56, 57]:
                for c_let in ["A", "B", "C", "D", "E", "F"]:
                    try:
                        get_writable_cell(ws, f"{c_let}{r}").value = None
                    except Exception:
                        pass

        elif sname == "ANA SAYFA KESİCİ":
            get_writable_cell(ws, "B74").value = "Test Tarihi"
            get_writable_cell(ws, "E74").value = ":"
            if test_date:
                s_val = date_to_excel_serial(test_date)
                cell_f74 = get_writable_cell(ws, "F74")
                cell_f74.value = s_val if s_val is not None else str(test_date)
                cell_f74.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B75").value = "Rapor Tarihi"
            get_writable_cell(ws, "E75").value = ":"
            if report_date:
                s_val = date_to_excel_serial(report_date)
                cell_f75 = get_writable_cell(ws, "F75")
                cell_f75.value = s_val if s_val is not None else str(report_date)
                cell_f75.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B76").value = "Unvan"
            get_writable_cell(ws, "E76").value = ":"
            get_writable_cell(ws, "F76").value = op_title

            for r in [77, 78]:
                for c_let in ["A", "B", "C", "D", "E", "F"]:
                    try:
                        get_writable_cell(ws, f"{c_let}{r}").value = None
                    except Exception:
                        pass

        elif "ANA SAYFA" in sname and sname != "ANA SAYFA KESİCİ":
            get_writable_cell(ws, "B79").value = "Test Tarihi"
            get_writable_cell(ws, "E79").value = ":"
            if test_date:
                s_val = date_to_excel_serial(test_date)
                cell_f79 = get_writable_cell(ws, "F79")
                cell_f79.value = s_val if s_val is not None else str(test_date)
                cell_f79.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B80").value = "Rapor Tarihi"
            get_writable_cell(ws, "E80").value = ":"
            if report_date:
                s_val = date_to_excel_serial(report_date)
                cell_f80 = get_writable_cell(ws, "F80")
                cell_f80.value = s_val if s_val is not None else str(report_date)
                cell_f80.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B81").value = "Unvan"
            get_writable_cell(ws, "E81").value = ":"
            get_writable_cell(ws, "F81").value = op_title

            for r in [82, 83, 84]:
                for c_let in ["A", "B", "C", "D", "E", "F"]:
                    try:
                        get_writable_cell(ws, f"{c_let}{r}").value = None
                    except Exception:
                        pass

        if sname != "KAPAK SAYFASI":
            for r in range(1, ws.max_row + 1):
                col_b_val = str(ws.cell(row=r, column=2).value or "").strip().upper()
                col_a_val = str(ws.cell(row=r, column=1).value or "").strip().upper()
                label_val = col_b_val if col_b_val else col_a_val
                if not label_val:
                    continue

                if label_val in ["UNVAN", "ÜNVAN"]:
                    target_cell = get_writable_cell(ws, f"F{r}")
                    target_cell.value = op_title
                elif label_val in ["TEST TARİHİ", "TEST TARIHI"]:
                    target_cell = get_writable_cell(ws, f"F{r}")
                    target_cell.value = "='KAPAK SAYFASI'!D53"
                    target_cell.number_format = "dd.mm.yyyy"
                elif any(kw in label_val for kw in ["ODA SİCİL NO", "ODA SICIL NO", "SICIL NO", "SİCİL NO", "EKİPNET NO", "EKIPNET NO"]):
                    for c_let in ["A", "B", "C", "D", "E", "F"]:
                        try:
                            get_writable_cell(ws, f"{c_let}{r}").value = None
                        except Exception:
                            pass


def fix_subpage_dates(wb, test_date: Optional[str] = None, report_date: Optional[str] = None):
    """
    Fixes sub-page OTURUM TARİHİ cells across all worksheets (TOPRAKLAMALAR, HV PF, LV PF, DİĞER, AKIM TRAFOLARI, etc.)
    by pointing them to 'KAPAK SAYFASI'!D53 formula and applying number_format 'dd.mm.yyyy'.
    Ensures date cell never evaluates to 0 or '00.01.1900'.
    """
    subpage_date_cells = {
        "OG SARGI MEVCUT KADEME": "J9",
        "AG SARGI": "J9",
        "İZOLASYON ": "J9",
        "Ç.O 34500": "J9",
        "TOPRAKLAMALAR": "J7",
        "HV PF": "J9",
        "LV PF": "J9",
        "KESİCİ İZOLASYON": "J8",
        "KESİCİ KONTAK": "J8",
        "AÇMA-KAPAMA": "J7",
        "DİĞER": "J7",
        "AKIM TRAFOLARI": "J7",
    }
    for sname in wb.sheetnames:
        ws = wb[sname]
        cell_ref = subpage_date_cells.get(sname)
        if cell_ref:
            try:
                cell = get_writable_cell(ws, cell_ref)
                cell.value = "='KAPAK SAYFASI'!D53"
                cell.number_format = "dd.mm.yyyy"
            except Exception:
                pass
        else:
            for r in range(1, 15):
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if val is not None and ("OTURUM TARİHİ" in str(val).upper() or "OTURUM TARIHI" in str(val).upper()):
                        target_col = c + 8 if c == 2 else c + 2
                        target_cell = ws.cell(row=r, column=target_col)
                        if type(target_cell).__name__ != 'MergedCell':
                            target_cell.value = "='KAPAK SAYFASI'!D53"
                            target_cell.number_format = "dd.mm.yyyy"


def process_oil_petition(wb, data_dict):
    """
    Dynamically replaces sample terms (date, brand, power_kva, serial_no)
    in the body paragraph (B10) of 'HERMETİK YAĞ DİLEKÇESİ' sheet without altering template layout or legal text.
    Also clears residual D16/D18 values on petition sheet.
    """
    for sname in wb.sheetnames:
        if "YAĞ DİLEKÇESİ" in sname.upper() or "YAG DILEKCESI" in sname.upper():
            ws = wb[sname]

            for cref in ["D16", "D18"]:
                try:
                    get_writable_cell(ws, cref).value = None
                except Exception:
                    pass

            try:
                c_name = data_dict.get("customer_name")
                if c_name:
                    get_writable_cell(ws, "J2").value = str(c_name).strip()
                else:
                    get_writable_cell(ws, "J2").value = "='KAPAK SAYFASI'!D9"

                addr = data_dict.get("address") or data_dict.get("location")
                if addr:
                    get_writable_cell(ws, "J3").value = str(addr).strip()
                else:
                    get_writable_cell(ws, "J3").value = "='KAPAK SAYFASI'!D11"
            except Exception:
                pass

            raw_date = data_dict.get("test_date") or data_dict.get("report_date")
            formatted_date = None
            if raw_date:
                s_d = str(raw_date).strip()
                if "." in s_d:
                    formatted_date = s_d
                elif "-" in s_d:
                    parts = s_d.split("-")
                    if len(parts) == 3:
                        formatted_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
            if not formatted_date:
                formatted_date = datetime.now().strftime("%d.%m.%Y")

            brand_val = str(data_dict.get("brand") or "").strip()
            if not brand_val:
                brand_val = "ASTOR"

            power_val = data_dict.get("power_kva")
            if power_val:
                str_p = str(power_val).strip()
                if not str_p.lower().endswith("kva"):
                    str_p += "kVA"
            else:
                str_p = "1000kVA"

            serial_val = str(data_dict.get("serial_no") or "").strip()
            if not serial_val:
                serial_val = "27-07840"

            canonical_text = (
                "\n              İşletmenizde bulunan trafoların testini {date} tarihinde yapmış bulunmaktayız.\n"
                "Test kapsamında istenilen trafo yağ analizi işleminin yapılabilmesi için trafodan yağ numunesi alınması gerekmektedir. "
                "İşletmenizde bulunan {brand} marka, {power_kva}, '{serial_no}' seri numaralı trafonun hermetik tip olmasından dolayı yağ numunesi alınmamıştır.          \n"
                "                                                                                                                                                               \n"
                "              Hermetik tip trafolar atmosfere kapalıdır ve numune alma yerleri mühürlüdür. Mühürlerin açılıp numune alınması durumunda trafonun hermetik özelliği bozulacağından dolayı, fabrika standartlarına göre arıza durumu olmadığı sürece hiçbir şekilde hermetik trafolardan yağ numunesi alınamaz.\n\n"
                "              \n"
                "              Bilgilerinize sunarız.\n\n"
            )

            cell_b10 = get_writable_cell(ws, "B10")
            current_val = cell_b10.value

            if current_val and isinstance(current_val, str) and ("işletmenizde bulunan" in current_val.lower() or "trafo" in current_val.lower()):
                text = current_val
                text = re.sub(r"\b\d{2}\.\d{2}\.\d{4}\b", formatted_date, text)
                if brand_val != "ASTOR":
                    text = text.replace("ASTOR", brand_val)
                if str_p != "1000kVA":
                    text = text.replace("1000kVA", str_p).replace("1000 kVA", str_p)
                if serial_val != "27-07840":
                    text = text.replace("27-07840", serial_val)
                cell_b10.value = text
            else:
                text = canonical_text.format(
                    date=formatted_date,
                    brand=brand_val,
                    power_kva=str_p,
                    serial_no=serial_val
                )
                cell_b10.value = text


def apply_format_and_column_width_fixes(wb):
    """
    Ensures date cells, GΩ 20°C calculated cells, and V DC [V] cells have explicit number_format
    and sufficient column width to eliminate '#####' display issues.
    """
    for sname in ["KAPAK SAYFASI", "ANA SAYFA"]:
        if sname in wb.sheetnames:
            ws = wb[sname]
            ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 14.0)
            for cref in ["D12", "D14", "D53", "D54"]:
                if cref in ws:
                    ws[cref].number_format = "dd.mm.yyyy"

    for sname in wb.sheetnames:
        norm_sname = sname.strip().upper()
        if norm_sname == "İZOLASYON":
            ws = wb[sname]
            for col_let in ["B", "D", "F", "H", "J", "L"]:
                ws.column_dimensions[col_let].width = max(ws.column_dimensions[col_let].width or 0, 13.0)
            for r in range(16, 32):
                for col_let in ["B", "D", "F", "H", "J", "L"]:
                    cell = ws[f"{col_let}{r}"]
                    if cell.value is not None:
                        if cell.number_format in ["General", "0.000000"]:
                            cell.number_format = "0.00"

    for sname in wb.sheetnames:
        if any(k in sname.upper() for k in ["AG SARGI", "OG SARGI"]):
            ws = wb[sname]
            ws.column_dimensions["G"].width = max(ws.column_dimensions["G"].width or 0, 15.0)
            for r in [24, 25, 26]:
                cell = ws[f"G{r}"]
                if cell.number_format in ["General", "0.000000"]:
                    cell.number_format = "0.00"


def resolve_default_template(mapped_type: str, backend_dir: str) -> str:
    """
    Resolves template file path in priority order for backend:
    1. backend/templates/hybrid/{hermetik_hybrid|gt_hybrid|kuru_tip_hybrid}.xlsx
    2. backend/templates/{HERMETİK TRAFO BAKIM RAPORU HİLMİ|KURU TİP HİLMİ|TR BAKIM RAPORU GT HİLMİ}.xlsx (legacy fallback)
    """
    type_filename_map = {
        "HERMETIK": ("hermetik_hybrid.xlsx", "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx"),
        "GT": ("gt_hybrid.xlsx", "TR BAKIM RAPORU GT HİLMİ.xlsx"),
        "KURU_TIP": ("kuru_tip_hybrid.xlsx", "KURU TİP HİLMİ.xlsx"),
    }

    hybrid_fname, old_fname = type_filename_map.get(mapped_type, type_filename_map["HERMETIK"])

    candidates = [
        os.path.abspath(os.path.join(backend_dir, "templates", "hybrid", hybrid_fname)),
        os.path.abspath(os.path.join(backend_dir, "templates", old_fname)),
    ]

    for cand in candidates:
        if os.path.exists(cand):
            return cand

    raise FileNotFoundError(f"No candidate template found for '{mapped_type}' in backend/templates! Checked paths: {candidates}")


def generate_excel_report(report: Report, photos: Optional[List[Photo]] = None, output_path: Optional[str] = None) -> str:
    mapped_type = (getattr(report, 'transformer_type', None) or getattr(report, 'report_type', None) or "HERMETIK").upper()
    if "KURU" in mapped_type:
        mapped_type = "KURU_TIP"
    elif "GT" in mapped_type or "TANK" in mapped_type:
        mapped_type = "GT"
    else:
        mapped_type = "HERMETIK"

    # __file__ is backend/app/services/excel_engine.py -> backend_dir is backend/
    backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = resolve_default_template(mapped_type, backend_dir)

    wb = openpyxl.load_workbook(template_path, data_only=False)

    data_dict = report.data_json or {}

    if mapped_type == "HERMETIK":
        data_dict["tank_mark_hermetik"] = "ü"
    elif mapped_type == "GT":
        data_dict["tank_mark_gt"] = "ü"
    elif mapped_type == "KURU_TIP":
        data_dict["tank_mark_kuru"] = "ü"

    if "address" not in data_dict or not data_dict["address"]:
        data_dict["address"] = data_dict.get("location", "")

    op_name = str(data_dict.get("operator_name") or data_dict.get("creator_display_name") or "").strip() or "Operatör"
    op_title = str(data_dict.get("operator_title") or data_dict.get("title") or "").strip() or "Elektrik Mühendisi"
    data_dict["operator_name"] = op_name
    data_dict["operator_title"] = op_title

    notes_text = data_dict.get("notes") or data_dict.get("notes_text") or ""
    if notes_text:
        notes_str = str(notes_text).strip()
        if not notes_str.upper().startswith("NOTLAR"):
            data_dict["notes"] = f"NOTLAR : {notes_str}"
        else:
            data_dict["notes"] = notes_str

    wr = data_dict.get("winding_resistance") if isinstance(data_dict.get("winding_resistance"), dict) else {}
    data_dict.setdefault("og_rab", wr.get("r_phase"))
    data_dict.setdefault("og_rbc", wr.get("s_phase"))
    data_dict.setdefault("og_rca", wr.get("t_phase"))

    gr = data_dict.get("grounding") if isinstance(data_dict.get("grounding"), dict) else {}
    data_dict.setdefault("ground_r_trafo_body", data_dict.get("ground_trafo_body") or gr.get("value"))

    # Insulation 30s/45s/60s fallback aliases
    data_dict.setdefault("iso_yg_tank_30s", data_dict.get("iso_og_gnd_30s") or data_dict.get("iso_og_gnd") or data_dict.get("iso_yg_tank"))
    data_dict.setdefault("iso_yg_tank_45s", data_dict.get("iso_og_gnd_45s"))
    data_dict.setdefault("iso_yg_tank_60s", data_dict.get("iso_og_gnd_60s"))

    data_dict.setdefault("iso_ag_tank_30s", data_dict.get("iso_ag_gnd_30s") or data_dict.get("iso_ag_gnd") or data_dict.get("iso_ag_tank"))
    data_dict.setdefault("iso_ag_tank_45s", data_dict.get("iso_ag_gnd_45s"))
    data_dict.setdefault("iso_ag_tank_60s", data_dict.get("iso_ag_gnd_60s"))

    data_dict.setdefault("iso_yg_ag_30s", data_dict.get("iso_og_ag_30s") or data_dict.get("iso_og_ag") or data_dict.get("iso_yg_ag"))
    data_dict.setdefault("iso_yg_ag_45s", data_dict.get("iso_og_ag_45s"))
    data_dict.setdefault("iso_yg_ag_60s", data_dict.get("iso_og_ag_60s"))

    cell_map_sheets = TYPE_CELL_MAPPINGS.get(mapped_type, TYPE_CELL_MAPPINGS["HERMETIK"])

    # 1. Pre-clear hardcoded sample data cells across all sheets
    pre_clear_sample_cells(wb)

    # 2. Main cell writing loop
    force_overwrite_keys = {
        "operator_name", "operator_title", "creator_display_name", "notes", "device_model", "device_serial"
    }

    for target_sheet_name, cell_map in cell_map_sheets.items():
        matched_sheet = None
        for sname in wb.sheetnames:
            if sname.strip() == target_sheet_name.strip():
                matched_sheet = sname
                break

        if matched_sheet:
            ws = wb[matched_sheet]
            if target_sheet_name.strip() in ["ANA SAYFA", "ANA SAYFA KESİCİ"]:
                process_checklist_pairs(ws, mapped_type, data_dict)

            for cell_ref, field_key in cell_map.items():
                if field_key.startswith("TODO_VERIFY"):
                    continue

                cell_obj = get_writable_cell(ws, cell_ref)

                if cell_obj.value is not None and str(cell_obj.value).strip().startswith("="):
                    if field_key not in force_overwrite_keys:
                        continue

                val = data_dict.get(field_key)
                if val is None or val == "" or (isinstance(val, list) and len(val) == 0):
                    cell_obj.value = None
                    continue

                if field_key.startswith("checklist_") or field_key.startswith("tank_mark_"):
                    if val is True or str(val).strip().lower() in ["true", "ü", "1", "evet"]:
                        cell_obj.value = "ü"
                    else:
                        cell_obj.value = None
                    continue

                if field_key in ["report_date", "test_date"]:
                    serial_val = date_to_excel_serial(val)
                    if serial_val is not None:
                        cell_obj.value = serial_val
                        cell_obj.number_format = "dd.mm.yyyy"
                    else:
                        cell_obj.value = None
                    continue

                if isinstance(val, bool):
                    cell_obj.value = "ü" if val else None
                elif isinstance(val, (int, float)):
                    cell_obj.value = val
                else:
                    str_val = str(val).strip()
                    if not str_val:
                        cell_obj.value = None
                    else:
                        clean_num = str_val.replace(',', '.')
                        try:
                            if '.' in clean_num:
                                cell_obj.value = float(clean_num)
                            else:
                                cell_obj.value = int(clean_num)
                        except Exception:
                            cell_obj.value = str_val

    # 3. Process signatures
    sig_path = data_dict.get("signature_path") or data_dict.get("signature")
    sheet_signature_map = SHEET_SIGNATURE_ANCHORS.get(mapped_type, SHEET_SIGNATURE_ANCHORS["HERMETIK"])

    for sname in wb.sheetnames:
        ws = wb[sname]
        target_anchor = None
        for k_sheet, a_cell in sheet_signature_map.items():
            if k_sheet.strip() == sname.strip():
                target_anchor = a_cell
                break
        process_sheet_signature(ws, target_anchor, sig_path)

    # 4. Process Kapak Photos
    if "KAPAK SAYFASI" in wb.sheetnames:
        process_kapak_photos(wb["KAPAK SAYFASI"], data_dict, photos)

    # 5. Clean legal text, personnel blocks, dates, oil petition
    clean_5070_text(wb)
    test_date_val = data_dict.get("test_date")
    report_date_val = data_dict.get("report_date")
    clean_left_personnel_block(wb, op_name, op_title, test_date=test_date_val, report_date=report_date_val, data_dict=data_dict)
    fix_subpage_dates(wb, test_date=test_date_val, report_date=report_date_val)
    process_oil_petition(wb, data_dict)

    # 6. Breaker pruning
    BREAKER_SHEETS = ["ANA SAYFA KESİCİ", "KESİCİ İZOLASYON", "KESİCİ KONTAK", "AÇMA-KAPAMA"]
    has_breaker = data_dict.get("has_breaker")
    if has_breaker is None:
        has_breaker = data_dict.get("breaker_included")
    if has_breaker is None:
        has_breaker = bool(data_dict.get("breaker_brand") or data_dict.get("breaker_iso_r_gnd") or data_dict.get("breaker_contact_r"))

    is_breaker_enabled = (has_breaker is True or str(has_breaker).strip().lower() in ["true", "1", "yes", "evet"])
    if not is_breaker_enabled:
        for b_sheet in BREAKER_SHEETS:
            if b_sheet in wb.sheetnames:
                del wb[b_sheet]

    # 7. Replace old sample names
    b_brand = str(data_dict.get("breaker_brand") or "").strip()
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    cval = cell.value
                    cval_u = cval.upper()
                    if "HİLMİ" in cval_u or "HILMI" in cval_u:
                        for old_term in ["Hilmi GÜL", "Hilmi GUL", "Hilmi"]:
                            cval = cval.replace(old_term, op_name)
                    if "ULUSOY" in cval_u:
                        cval = cval.replace("ULUSOY", b_brand)
                        if "ULusoy" in cval:
                            cval = cval.replace("ULusoy", b_brand)
                    cell.value = cval

    # 8. Format & column width fixes
    apply_format_and_column_width_fixes(wb)

    # 9. Save & fix rels
    if not output_path:
        os.makedirs(settings.EXPORT_DIR, exist_ok=True)
        report_id = getattr(report, "id", "temp")
        output_path = os.path.join(settings.EXPORT_DIR, f"report_{report_id}.xlsx")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    fix_xlsx_rels(output_path)
    return output_path


def generate_report_excel(report: Report, photos: Optional[List[Photo]] = None, signature_path: Optional[str] = None, output_path: Optional[str] = None) -> str:
    """
    Compatibility wrapper matching backend API expected signature.
    """
    if photos is None:
        photos = []

    if signature_path and isinstance(report.data_json, dict) and "signature_path" not in report.data_json:
        report.data_json["signature_path"] = signature_path

    return generate_excel_report(report, photos, output_path)


def fix_xlsx_rels(xlsx_path: str):
    """
    Fixes relationship target paths in .xlsx zip archive by converting
    absolute targets starting with '/xl/' (e.g. '/xl/drawings/drawing1.xml', '/xl/media/image1.png')
    to standard relative targets without leading '/xl/'.
    """
    try:
        buffer = io.BytesIO()
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    content = zin.read(item.filename)
                    if item.filename.endswith('.rels'):
                        try:
                            root = ET.fromstring(content)
                            modified = False
                            parts = item.filename.split('/')
                            if len(parts) >= 2 and parts[-2] == '_rels':
                                base_dir = '/'.join(parts[:-2])
                            else:
                                base_dir = '/'.join(parts[:-1])

                            for child in root:
                                tgt = child.get('Target')
                                if tgt and tgt.startswith('/xl/'):
                                    full_target = tgt[1:]
                                    rel_target = os.path.relpath(full_target, base_dir).replace('\\', '/')
                                    child.set('Target', rel_target)
                                    modified = True

                            if modified:
                                content = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                        except Exception:
                            pass
                    zout.writestr(item, content)

        with open(xlsx_path, 'wb') as f:
            f.write(buffer.getvalue())
    except Exception as e:
        sys.stderr.write(f"Warning: Failed to fix rels targets in '{xlsx_path}': {e}\n")
