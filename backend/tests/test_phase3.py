import os
import pytest
import openpyxl
from fastapi.testclient import TestClient

os.environ["DATABASE_URL"] = "sqlite:///./test_traforeport_phase3.db"

from app.main import app
from app.db.session import engine, Base
from create_initial_admin import seed_data

client = TestClient(app)

@pytest.fixture(scope="session", autouse=True)
def setup_test_db():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    seed_data()
    yield
    Base.metadata.drop_all(bind=engine)
    if os.path.exists("./test_traforeport_phase3.db"):
        try:
            os.remove("./test_traforeport_phase3.db")
        except PermissionError:
            pass

def get_auth_token(identifier, password):
    resp = client.post("/api/v1/auth/login", json={"identifier": identifier, "password": password})
    return resp.json()["access_token"]

@pytest.mark.parametrize("report_type", ["HERMETIK", "KURU_TIP", "GT"])
def test_excel_generation_and_download_all_templates(report_type):
    # 1. Admin login & generate fresh invite code
    admin_token = get_auth_token("05000000000", "Admin123!")
    admin_headers = {"Authorization": f"Bearer {admin_token}"}
    code_resp = client.post("/api/v1/admin/codes", json={}, headers=admin_headers)
    fresh_code = code_resp.json()["code"]

    # 2. Register employee using fresh invite code
    emp_phone = f"0544{report_type[:4].ljust(7, '0')}"[:11]
    client.post("/api/v1/auth/register", json={
        "full_name": f"Teknisyen {report_type}",
        "phone": emp_phone,
        "invite_code": fresh_code,
        "password": "Password123"
    })

    token = get_auth_token(emp_phone, "Password123")
    headers = {"Authorization": f"Bearer {token}"}

    # 3. Create report
    create_resp = client.post("/api/v1/reports", json={
        "title": f"Musteri {report_type} - TR1 - 30.07.2026",
        "report_type": report_type,
        "maintenance_type": "maintenance",
        "status": "draft",
        "customer_name": f"Musteri {report_type}",
        "trafo_label": "TR1",
        "test_date": "2026-07-30",
        "report_date": "2026-07-30",
        "data_json": {
            "brand": "Schneider",
            "power_kva": 1600,
            "og_r_a": 3.25,
            "og_r_b": 3.28,
            "og_r_c": 3.26
        }
    }, headers=headers)
    assert create_resp.status_code == 201
    report_id = create_resp.json()["id"]

    # 4. Finalize report (Generate Excel)
    fin_resp = client.post(f"/api/v1/reports/{report_id}/finalize", headers=headers)
    assert fin_resp.status_code == 200
    fin_data = fin_resp.json()
    assert fin_data["status"] == "final"
    assert fin_data["excel_path"] is not None
    assert os.path.exists(fin_data["excel_path"])

    # 5. Validate generated Excel contents with openpyxl
    wb = openpyxl.load_workbook(fin_data["excel_path"], data_only=True)
    assert "KAPAK SAYFASI" in wb.sheetnames
    assert "ANA SAYFA" in wb.sheetnames
    
    ws_kapak = wb["KAPAK SAYFASI"]
    assert ws_kapak["D9"].value == f"Musteri {report_type}"
    assert ws_kapak["D10"].value == "TR1"
    
    date_val = ws_kapak["D12"].value
    # openpyxl converts date-formatted cells to datetime/date object or numeric serial number
    if hasattr(date_val, "year"):
        assert date_val.year == 2026 and date_val.month == 7 and date_val.day == 30
    else:
        assert float(date_val) == 46233.0

    # 6. Test download endpoint
    dl_resp = client.get(f"/api/v1/reports/{report_id}/download", headers=headers)
    assert dl_resp.status_code == 200
    assert dl_resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(dl_resp.content) > 1000

