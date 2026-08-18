import os
import openpyxl
import zipfile
import xml.etree.ElementTree as ET

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TEMPLATES_DIR = os.path.join(REPO_ROOT, "backend", "templates")
HYBRID_DIR = os.path.join(TEMPLATES_DIR, "hybrid")
TOOLS_DIR = os.path.join(REPO_ROOT, "tools")

SRC_PATH = os.path.join(TEMPLATES_DIR, "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx")
HYBRID_PATH = os.path.join(HYBRID_DIR, "hermetik_hybrid.xlsx")
OUTPUT_PATH = os.path.join(TOOLS_DIR, "t1_minimal_output.xlsx")

def inspect_sheet_properties(wb, sheet_name):
    matched = None
    for name in wb.sheetnames:
        if name.strip() == sheet_name.strip():
            matched = name
            break
    if not matched:
        return None
    ws = wb[matched]
    
    props = {}
    
    # 1. sheet_format
    sf = getattr(ws, 'sheet_format', None)
    props['defaultRowHeight'] = getattr(sf, 'defaultRowHeight', None)
    props['defaultColWidth'] = getattr(sf, 'defaultColWidth', None)
    props['customHeight'] = getattr(sf, 'customHeight', None)
    props['zeroHeight'] = getattr(sf, 'zeroHeight', None)
    
    # 2. sheet_view
    sv = getattr(ws, 'sheet_view', None)
    props['view_type'] = getattr(sv, 'view', None)
    props['zoomScale'] = getattr(sv, 'zoomScale', None)
    props['zoomScaleNormal'] = getattr(sv, 'zoomScaleNormal', None)
    props['showGridLines'] = getattr(sv, 'showGridLines', None)
    
    # 3. page_setup
    ps = getattr(ws, 'page_setup', None)
    props['orientation'] = getattr(ps, 'orientation', None)
    props['paperSize'] = getattr(ps, 'paperSize', None)
    props['scale'] = getattr(ps, 'scale', None)
    props['fitToWidth'] = getattr(ps, 'fitToWidth', None)
    props['fitToHeight'] = getattr(ps, 'fitToHeight', None)
    
    # 4. pageSetUpPr
    spr = getattr(ws, 'sheet_properties', None)
    pspr = getattr(spr, 'pageSetUpPr', None) if spr else None
    props['fitToPage'] = getattr(pspr, 'fitToPage', None) if pspr else None
    props['autoPageBreaks'] = getattr(pspr, 'autoPageBreaks', None) if pspr else None
    
    # 5. page_margins
    pm = getattr(ws, 'page_margins', None)
    props['margins'] = {
        'top': getattr(pm, 'top', None),
        'bottom': getattr(pm, 'bottom', None),
        'left': getattr(pm, 'left', None),
        'right': getattr(pm, 'right', None),
        'header': getattr(pm, 'header', None),
        'footer': getattr(pm, 'footer', None),
    } if pm else {}
    
    # 6. row heights sample
    row_heights = {}
    for r in range(1, min(85, ws.max_row + 1)):
        rdim = ws.row_dimensions.get(r)
        h = rdim.height if rdim else None
        ch = rdim.customHeight if rdim else None
        row_heights[r] = (h, ch)
    props['row_heights_sample'] = row_heights
    
    # 7. column widths sample
    col_widths = {}
    for c_let, cdim in ws.column_dimensions.items():
        col_widths[c_let] = (cdim.width, cdim.bestFit, cdim.customWidth if hasattr(cdim, 'customWidth') else None)
    props['col_widths'] = col_widths
    
    return props

def inspect_zip_xml(xlsx_path, target_sheet_title):
    sheet_xml_map = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # Read workbook.xml to get sheet rId -> name
        wb_xml = z.read('xl/workbook.xml')
        wb_tree = ET.fromstring(wb_xml)
        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
              'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        
        rel_map = {}
        wb_rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_tree = ET.fromstring(wb_rels_xml)
        for child in rels_tree:
            rid = child.get('Id')
            target = child.get('Target')
            rel_map[rid] = target

        for s in wb_tree.findall('.//ns:sheet', ns):
            s_name = s.get('name')
            s_rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            target_path = rel_map.get(s_rid, '')
            if target_path.startswith('/xl/'):
                target_path = target_path[4:]
            elif not target_path.startswith('xl/'):
                target_path = 'xl/' + target_path
            sheet_xml_map[s_name.strip()] = target_path

        xml_path = sheet_xml_map.get(target_sheet_title.strip())
        if not xml_path or xml_path not in z.namelist():
            return None
        
        xml_content = z.read(xml_path)
        tree = ET.fromstring(xml_content)
        
        info = {}
        # sheetView
        sv = tree.find('.//ns:sheetView', ns)
        if sv is not None:
            info['view'] = sv.get('view')
            info['zoomScale'] = sv.get('zoomScale')
            info['zoomScaleNormal'] = sv.get('zoomScaleNormal')
            info['zoomScalePageLayoutView'] = sv.get('zoomScalePageLayoutView')
            info['workbookViewId'] = sv.get('workbookViewId')
            
        # sheetFormatPr
        sf = tree.find('.//ns:sheetFormatPr', ns)
        if sf is not None:
            info['sf_defaultRowHeight'] = sf.get('defaultRowHeight')
            info['sf_customHeight'] = sf.get('customHeight')
            info['sf_zeroHeight'] = sf.get('zeroHeight')
            
        # pageSetup
        ps = tree.find('.//ns:pageSetup', ns)
        if ps is not None:
            info['ps_orientation'] = ps.get('orientation')
            info['ps_paperSize'] = ps.get('paperSize')
            info['ps_scale'] = ps.get('scale')
            info['ps_fitToWidth'] = ps.get('fitToWidth')
            info['ps_fitToHeight'] = ps.get('fitToHeight')
            
        # pageSetUpPr
        pspr = tree.find('.//ns:pageSetUpPr', ns)
        if pspr is not None:
            info['pspr_fitToPage'] = pspr.get('fitToPage')
            info['pspr_autoPageBreaks'] = pspr.get('autoPageBreaks')

        return info

def main():
    print("Loading Workbooks...")
    wb_src = openpyxl.load_workbook(SRC_PATH, data_only=False)
    wb_hyb = openpyxl.load_workbook(HYBRID_PATH, data_only=False)
    wb_out = openpyxl.load_workbook(OUTPUT_PATH, data_only=False)
    
    target_sheets = ["KAPAK SAYFASI", "ANA SAYFA", "ANA SAYFA KESİCİ", "KESİCİ İZOLASYON", "KESİCİ KONTAK", "AÇMA-KAPAMA", "OG SARGI MEVCUT KADEME"]
    
    for sname in target_sheets:
        print(f"\n==================================================")
        print(f"SHEET: {sname}")
        print(f"==================================================")
        
        p_src = inspect_sheet_properties(wb_src, sname)
        p_hyb = inspect_sheet_properties(wb_hyb, sname)
        p_out = inspect_sheet_properties(wb_out, sname)
        
        xml_src = inspect_zip_xml(SRC_PATH, sname)
        xml_hyb = inspect_zip_xml(HYBRID_PATH, sname)
        xml_out = inspect_zip_xml(OUTPUT_PATH, sname)
        
        print("\n--- 1. OPENPYXL PROPERTIES ---")
        if p_src:
            print(f"Src: defRH={p_src['defaultRowHeight']} | view={p_src['view_type']} | zoom={p_src['zoomScale']} | orient={p_src['orientation']} | scale={p_src['scale']} | fitW={p_src['fitToWidth']} | fitH={p_src['fitToHeight']} | fitPage={p_src['fitToPage']}")
        if p_hyb:
            print(f"Hyb: defRH={p_hyb['defaultRowHeight']} | view={p_hyb['view_type']} | zoom={p_hyb['zoomScale']} | orient={p_hyb['orientation']} | scale={p_hyb['scale']} | fitW={p_hyb['fitToWidth']} | fitH={p_hyb['fitToHeight']} | fitPage={p_hyb['fitToPage']}")
        if p_out:
            print(f"Out: defRH={p_out['defaultRowHeight']} | view={p_out['view_type']} | zoom={p_out['zoomScale']} | orient={p_out['orientation']} | scale={p_out['scale']} | fitW={p_out['fitToWidth']} | fitH={p_out['fitToHeight']} | fitPage={p_out['fitToPage']}")
        
        print("\n--- 2. RAW XML STRUCTURE (sheetView / pageSetup / sheetFormatPr) ---")
        print(f"Src XML: {xml_src}")
        print(f"Hyb XML: {xml_hyb}")
        print(f"Out XML: {xml_out}")

if __name__ == "__main__":
    main()
