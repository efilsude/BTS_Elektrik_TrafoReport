import os
import sys
import openpyxl
import zipfile
import xml.etree.ElementTree as ET

sys.stdout.reconfigure(encoding='utf-8')

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TEMPLATES_DIR = os.path.join(REPO_ROOT, "backend", "templates")
HILMI_PATH = os.path.join(TEMPLATES_DIR, "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx")
GT_PATH = os.path.join(TEMPLATES_DIR, "TR BAKIM RAPORU GT HİLMİ.xlsx")
KURU_PATH = os.path.join(TEMPLATES_DIR, "KURU TİP HİLMİ.xlsx")
REF_PATH = os.path.join(TEMPLATES_DIR, "reference", "yeni_tasarim_referans.xlsx")
HYBRID_PATH = os.path.join(TEMPLATES_DIR, "hybrid", "hermetik_hybrid.xlsx")
OUTPUT_PATH = os.path.join(REPO_ROOT, "tools", "t1_minimal_output.xlsx")

def inspect_ana_sayfa_kesici_cells():
    wb = openpyxl.load_workbook(HILMI_PATH, data_only=False)
    if "ANA SAYFA KESİCİ" not in wb.sheetnames:
        print("ANA SAYFA KESİCİ not in HİLMİ.xlsx")
        return
    ws = wb["ANA SAYFA KESİCİ"]
    print("\n==================================================")
    print("ANALYZING ALL CELLS IN 'ANA SAYFA KESİCİ'")
    print("==================================================")
    
    non_empty = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                cref = openpyxl.utils.get_column_letter(c) + str(r)
                sval = str(val).strip()
                is_formula = sval.startswith("=")
                non_empty.append((cref, r, c, sval, is_formula))
                
    print(f"Total non-empty cells in ANA SAYFA KESİCİ: {len(non_empty)}")
    
    print("\n--- NON-FORMULA VALUES (VARIABLE vs FIXED CANDIDATES) ---")
    for cref, r, c, sval, is_formula in non_empty:
        if not is_formula and r <= 85:
            print(f"  {cref:5s} (Row {r:2d}, Col {c:2d}): {sval!r}")

def inspect_personnel_blocks_all_sheets(xlsx_path, tag):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    print(f"\n==================================================")
    print(f"PERSONNEL / DATE BLOCKS IN ALL SHEETS ({tag}: {os.path.basename(xlsx_path)})")
    print(f"==================================================")
    
    keywords = ["TEST TARİHİ", "RAPOR TARİHİ", "UNVAN", "ÜNVAN", "SİCİL", "SICIL", "EKİPNET", "EKIPNET", "ONAYLAYAN", "İSİM SOYAD", "ISIM SOYAD"]
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        matches = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    sval = str(val).strip().upper()
                    if any(kw in sval for kw in keywords):
                        cref = openpyxl.utils.get_column_letter(c) + str(r)
                        val_cell = ws.cell(row=r, column=c+1).value if c+1 <= ws.max_column else None
                        val_cell_f = ws.cell(row=r, column=6).value  # Col F
                        matches.append((cref, ws.cell(row=r, column=c).value, val_cell, val_cell_f))
        if matches:
            print(f"\nSheet '{sname}':")
            for m in matches:
                print(f"  Label cell {m[0]}: {m[1]!r} | next cell: {m[2]!r} | Col F cell: {m[3]!r}")

def compare_vertical_layout_props(hilmi_path, hybrid_path, output_path):
    print(f"\n==================================================")
    print("COMPARING VERTICAL LAYOUT PROPS (HİLMİ vs HYBRID vs OUTPUT)")
    print("==================================================")
    
    target_sheets = ["KAPAK SAYFASI", "ANA SAYFA", "ANA SAYFA KESİCİ", "KESİCİ İZOLASYON", "KESİCİ KONTAK", "AÇMA-KAPAMA"]
    
    for sname in target_sheets:
        print(f"\n--- SHEET: {sname} ---")
        for tag, path in [("ORIGINAL HİLMİ", hilmi_path), ("HYBRID MASTER", hybrid_path), ("GENERATED OUTPUT", output_path)]:
            if not os.path.exists(path):
                continue
            with zipfile.ZipFile(path, 'r') as z:
                wb_tree = ET.fromstring(z.read('xl/workbook.xml'))
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                      'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
                rel_map = {c.get('Id'): c.get('Target') for c in ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))}
                
                matched_target = None
                for s in wb_tree.findall('.//ns:sheet', ns):
                    if s.get('name').strip() == sname.strip():
                        matched_target = rel_map.get(s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'))
                        break
                if not matched_target:
                    print(f"  {tag:16s}: Sheet not found")
                    continue
                if matched_target.startswith('/xl/'):
                    matched_target = matched_target[4:]
                elif not matched_target.startswith('xl/'):
                    matched_target = 'xl/' + matched_target
                    
                sheet_xml = z.read(matched_target)
                tree = ET.fromstring(sheet_xml)
                
                sv = tree.find('.//ns:sheetView', ns)
                ps = tree.find('.//ns:pageSetup', ns)
                pspr = tree.find('.//ns:pageSetUpPr', ns)
                pm = tree.find('.//ns:pageMargins', ns)
                sf = tree.find('.//ns:sheetFormatPr', ns)
                
                sv_view = sv.get('view') if sv is not None else None
                zoom = sv.get('zoomScale') if sv is not None else None
                zoom_norm = sv.get('zoomScaleNormal') if sv is not None else None
                
                ps_scale = ps.get('scale') if ps is not None else None
                fit_w = ps.get('fitToWidth') if ps is not None else None
                fit_h = ps.get('fitToHeight') if ps is not None else None
                paper = ps.get('paperSize') if ps is not None else None
                orient = ps.get('orientation') if ps is not None else None
                
                fit_page = pspr.get('fitToPage') if pspr is not None else None
                
                def_rh = sf.get('defaultRowHeight') if sf is not None else None
                
                m_top = pm.get('top') if pm is not None else None
                m_bot = pm.get('bottom') if pm is not None else None
                
                print(f"  {tag:16s}: view={sv_view} | zoom={zoom} | zoomNorm={zoom_norm} | ps_scale={ps_scale} | fitPage={fit_page} | fitW={fit_w} | fitH={fit_h} | paper={paper} | orient={orient} | defRH={def_rh} | m_top={m_top} | m_bot={m_bot}")

def main():
    inspect_ana_sayfa_kesici_cells()
    inspect_personnel_blocks_all_sheets(HILMI_PATH, "ORIGINAL HİLMİ")
    compare_vertical_layout_props(HILMI_PATH, HYBRID_PATH, OUTPUT_PATH)

if __name__ == "__main__":
    main()
