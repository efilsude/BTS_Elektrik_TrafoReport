import os
import sys
import json
import openpyxl

# Add backend directory to sys.path
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
BACKEND_DIR = os.path.join(REPO_ROOT, "backend")
sys.path.insert(0, BACKEND_DIR)

from app.services.excel_engine import generate_excel_report

class MockReport:
    def __init__(self, report_type, data_dict):
        self.report_type = report_type
        self.transformer_type = report_type
        self.customer_name = data_dict.get("customer_name", "MOCK_MUSTERI")
        self.trafo_label = data_dict.get("trafo_label", "TR-MOCK")
        self.data_json = data_dict

def test_backend_generation():
    types_to_test = ["HERMETIK", "KURU_TIP", "GT"]
    
    sample_data = {
        "customer_name": "ACME CORP INC",
        "trafo_label": "TR-01-MAIN",
        "address": "OSB 3. Cadde No: 45",
        "test_date": "18.08.2026",
        "report_date": "18.08.2026",
        "operator_name": "Mehmet Demir",
        "operator_title": "Elektrik Başmühendisi",
        "brand": "SCHNEIDER",
        "power_kva": "1600kVA",
        "serial_no": "SN-2026-9999",
        "has_breaker": True,
        "breaker_brand": "SIEMENS",
        "breaker_model": "3AH5",
        "breaker_serial_no": "SIE-887766",
        "breaker_year": "2024",
        "breaker_voltage": "36000",
        "breaker_motor_voltage": "220V DC",
        "breaker_coil_voltage": "220V DC",
        "breaker_control_visual": True,
        "breaker_control_cleanliness": False,
        "breaker_iso_r_gnd": "20000",
        "breaker_contact_r": "18",
        "breaker_notes": "Backend test basariyla yazildi.",
        "checklist_1": True,
        "checklist_2": False,
    }

    all_passed = True

    for rtype in types_to_test:
        print(f"\n--- Testing Backend excel_engine for {rtype} ---")
        output_path = os.path.join(REPO_ROOT, "tools", f"backend_test_{rtype.lower()}_output.xlsx")
        report = MockReport(rtype, sample_data)
        
        generated_file = generate_excel_report(report, photos=[], output_path=output_path)
        print(f"Generated file: {generated_file}")
        
        wb = openpyxl.load_workbook(generated_file, data_only=False)
        sheet_count = len(wb.sheetnames)
        print(f"Sheet count: {sheet_count}")
        
        # Verify 16 sheets for hybrid template
        expected_sheets = 13 if rtype == "GT" else 16
        if sheet_count != expected_sheets:
            print(f"FAIL: Expected {expected_sheets} sheets for {rtype}, got {sheet_count}")
            all_passed = False
            
        # Verify Kapak customer name
        ws_kapak = wb["KAPAK SAYFASI"]
        cust_name_val = ws_kapak["D9"].value
        print(f"KAPAK D9 (Customer Name): {cust_name_val!r}")
        if cust_name_val != "ACME CORP INC":
            print(f"FAIL: Customer name mismatch")
            all_passed = False
            
        # Verify Breaker Sheet values
        ws_breaker = wb["ANA SAYFA KESİCİ"]
        b_brand = ws_breaker["G11"].value
        b_motor_v = ws_breaker["G17"].value
        b_vis_evet = ws_breaker["G24"].value
        b_vis_hayir = ws_breaker["I24"].value
        b_clean_evet = ws_breaker["P24"].value
        b_clean_hayir = ws_breaker["R24"].value
        b_iso_r = ws_breaker["G50"].value
        b_notes = ws_breaker["D68"].value
        
        print(f"KESİCİ G11 (Brand): {b_brand!r}")
        print(f"KESİCİ G17 (Motor V): {b_motor_v!r}")
        print(f"KESİCİ G24/I24 (Visual): Evet={b_vis_evet!r}, Hayir={b_vis_hayir!r}")
        print(f"KESİCİ P24/R24 (Clean): Evet={b_clean_evet!r}, Hayir={b_clean_hayir!r}")
        print(f"KESİCİ G50 (Iso R): {b_iso_r!r}")
        print(f"KESİCİ D68 (Notes): {b_notes!r}")
        
        # Verify residue cleanup
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    if val is not None and isinstance(val, str):
                        uval = val.upper()
                        if "OK AMBALAJ" in uval:
                            print(f"FAIL: Residual 'OK AMBALAJ' found in {sname} cell ({r},{c})")
                            all_passed = False
                        if "HİLMİ GÜL" in uval or "HILMI GUL" in uval:
                            print(f"FAIL: Residual 'Hilmi GÜL' found in {sname} cell ({r},{c})")
                            all_passed = False
                        if "ULUSOY" in uval:
                            print(f"FAIL: Residual 'ULUSOY' found in {sname} cell ({r},{c})")
                            all_passed = False

    print(f"\n==================================================")
    print(f"BACKEND EXCEL ENGINE TEST ALL PASSED: {all_passed}")
    print(f"==================================================")
    return all_passed

if __name__ == "__main__":
    test_backend_generation()
