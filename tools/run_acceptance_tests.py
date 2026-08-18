import json
import os
import subprocess
import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TOOLS_DIR = os.path.join(REPO_ROOT, "tools")

def run_t1():
    print("\n--- RUNNING T1: Minimal JSON Test ---")
    json_path = os.path.join(TOOLS_DIR, "t1_minimal.json")
    out_path = os.path.join(TOOLS_DIR, "t1_minimal_output.xlsx")
    
    t1_payload = {
        "customer_name": "MINIMAL_MUSTERI",
        "trafo_label": "TR-MINIMAL",
        "transformerType": "hermetik",
        "has_breaker": True,
        "breaker_included": True,
        "test_date": "17.08.2026",
        "report_date": "17.08.2026",
        "operator_name": "Ahmet Yilmaz",
        "operator_title": "Elektrik Muhendisi"
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(t1_payload, f, indent=2)
        
    cmd = [
        "python", os.path.join(TOOLS_DIR, "generate_excel.py"),
        "--json", json_path,
        "--template-type", "hermetik",
        "--has-breaker", "true",
        "--output", out_path
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    print(f"CLI Result: exit_code={res.returncode}")
    if res.returncode != 0:
        print(f"CLI stderr: {res.stderr}")
        return False
        
    # Openpyxl inspection on T1 output
    wb = openpyxl.load_workbook(out_path, data_only=False)
    ws = wb["ANA SAYFA KESİCİ"]
    
    ulusoy_count = 0
    hilmi_count = 0
    non_formula_val_count = 0
    
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                sval = str(val).strip()
                if "ULUSOY" in sval.upper():
                    ulusoy_count += 1
                if "HİLMİ" in sval.upper() or "HILMI" in sval.upper():
                    hilmi_count += 1
                if not sval.startswith("="):
                    non_formula_val_count += 1
                    
    # Check specific sample cells
    g11_val = ws["G11"].value
    o11_val = ws["O11"].value
    i24_val = ws["I24"].value
    c50_val = ws["C50"].value
    
    # Assert FIXED headers and labels remain intact (not deleted)
    b9_hdr = str(ws["B9"].value or "").strip()
    b21_hdr = str(ws["B21"].value or "").strip()
    i22_lbl = str(ws["I22"].value or "").strip()
    j22_lbl = str(ws["J22"].value or "").strip()
    b39_lbl = str(ws["B39"].value or "").strip()
    f50_unit = str(ws["F50"].value or "").strip()
    k74_hdr = str(ws["K74"].value or "").strip()
    
    print(f"[T1] ULUSOY Count in 'ANA SAYFA KESİCİ': {ulusoy_count}")
    print(f"[T1] HİLMİ Count in 'ANA SAYFA KESİCİ': {hilmi_count}")
    print(f"[T1] Sample Cell G11 (Marka): {g11_val!r}")
    print(f"[T1] Sample Cell O11 (Tip): {o11_val!r}")
    print(f"[T1] Sample Cell I24 (Tik): {i24_val!r}")
    print(f"[T1] Sample Cell C50 (Test Volt): {c50_val!r}")
    print(f"[T1] FIXED Header B9: {b9_hdr!r}")
    print(f"[T1] FIXED Header B21: {b21_hdr!r}")
    print(f"[T1] FIXED Label I22/J22: {i22_lbl!r} / {j22_lbl!r}")
    print(f"[T1] FIXED Label B39: {b39_lbl!r}")
    print(f"[T1] FIXED Unit F50: {f50_unit!r}")
    print(f"[T1] FIXED Header K74: {k74_hdr!r}")
    
    fixed_ok = ("KESİCİ" in b9_hdr) and ("KONTROL" in b21_hdr) and (i22_lbl == "Evet") and (b39_lbl == "L-1") and (f50_unit == "kV") and ("ONAYLAYAN" in k74_hdr)
    t1_success = (ulusoy_count == 0) and (g11_val is None) and (o11_val is None) and (i24_val is None) and (c50_val is None) and fixed_ok
    print(f"[T1] PASS: {t1_success}")
    return t1_success

def run_t2():
    print("\n--- RUNNING T2: Full Breaker + Partial Measurements Test ---")
    json_path = os.path.join(TOOLS_DIR, "t2_breaker.json")
    out_path = os.path.join(TOOLS_DIR, "t2_breaker_output.xlsx")
    
    t2_payload = {
        "customer_name": "TEST_MUSTERI_2",
        "trafo_label": "TR-TEST-2",
        "transformerType": "hermetik",
        "has_breaker": True,
        "breaker_included": True,
        "test_date": "17.08.2026",
        "report_date": "17.08.2026",
        "operator_name": "Ahmet Yilmaz",
        "operator_title": "Elektrik Muhendisi",
        "breaker_brand": "ABB",
        "breaker_model": "VD4",
        "breaker_serial_no": "ABB-98765",
        "breaker_year": "2023",
        "breaker_rated_current": "630",
        "breaker_voltage": "36000",
        "breaker_motor_voltage": "220V DC",
        "breaker_coil_voltage": "220V DC",
        "breaker_control_visual": True,
        "breaker_control_cleanliness": False,
        "breaker_notes": "Kesici test ve kontrolu yapildi.",
        "breaker_iso_r_gnd": "15000",
        "breaker_contact_r": "25"
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(t2_payload, f, indent=2)
        
    cmd = [
        "python", os.path.join(TOOLS_DIR, "generate_excel.py"),
        "--json", json_path,
        "--template-type", "hermetik",
        "--has-breaker", "true",
        "--output", out_path
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    print(f"CLI Result: exit_code={res.returncode}")
    if res.returncode != 0:
        print(f"CLI stderr: {res.stderr}")
        return False
        
    wb = openpyxl.load_workbook(out_path, data_only=False)
    ws = wb["ANA SAYFA KESİCİ"]
    
    g11_val = ws["G11"].value  # breaker_brand
    g13_val = ws["G13"].value  # breaker_model
    o11_val = ws["O11"].value  # breaker_serial_no
    g15_val = ws["G15"].value  # breaker_voltage
    g17_val = ws["G17"].value  # breaker_motor_voltage
    o17_val = ws["O17"].value  # breaker_coil_voltage
    g24_val = ws["G24"].value  # breaker_control_visual Evet ("ü")
    i24_val = ws["I24"].value  # breaker_control_visual Hayır (None)
    p24_val = ws["P24"].value  # breaker_control_cleanliness Evet (None)
    r24_val = ws["R24"].value  # breaker_control_cleanliness Hayır ("ü")
    g50_val = ws["G50"].value  # breaker_iso_r_gnd
    c39_val = ws["C39"].value  # unprovided step voltage (should be None)
    d68_val = ws["D68"].value  # breaker_notes
    f74_val = ws["F74"].value  # Test Tarihi (serial date)
    f75_val = ws["F75"].value  # Rapor Tarihi (serial date)
    
    print(f"[T2] G11 (Brand): {g11_val!r} (expected 'ABB')")
    print(f"[T2] G13 (Model): {g13_val!r} (expected 'VD4')")
    print(f"[T2] O11 (Serial): {o11_val!r} (expected 'ABB-98765')")
    print(f"[T2] G15 (Volt): {g15_val!r} (expected 36000)")
    print(f"[T2] G17 (Motor V): {g17_val!r} (expected '220V DC')")
    print(f"[T2] O17 (Coil V): {o17_val!r} (expected '220V DC')")
    print(f"[T2] G24/I24 (Visual Ctrl): Evet={g24_val!r}/Hayir={i24_val!r} (expected 'ü'/None)")
    print(f"[T2] P24/R24 (Clean Ctrl): Evet={p24_val!r}/Hayir={r24_val!r} (expected None/'ü')")
    print(f"[T2] G50 (Iso R): {g50_val!r} (expected 15000)")
    print(f"[T2] D68 (Notes): {d68_val!r} (expected 'Kesici test ve kontrolu yapildi.')")
    print(f"[T2] F74/F75 (Dates): Test={f74_val!r}, Rapor={f75_val!r}")
    print(f"[T2] C39 (Unprovided Step Volt): {c39_val!r} (expected None)")
    
    t2_success = (
        (g11_val == "ABB") and 
        (g13_val == "VD4") and 
        (g15_val == 36000) and
        (g17_val == "220V DC") and
        (g24_val == "ü") and
        (i24_val is None) and
        (r24_val == "ü") and
        (p24_val is None) and
        (c39_val is None) and
        (f74_val is not None)
    )
    print(f"[T2] PASS: {t2_success}")
    return t2_success

def run_t3():
    print("\n--- RUNNING T3: has_breaker=false Pruning Test ---")
    json_path = os.path.join(TOOLS_DIR, "t3_no_breaker.json")
    out_path = os.path.join(TOOLS_DIR, "t3_no_breaker_output.xlsx")
    
    t3_payload = {
        "customer_name": "MINIMAL_NO_BREAKER",
        "trafo_label": "TR-NO-BREAKER",
        "transformerType": "hermetik",
        "has_breaker": False,
        "breaker_included": False,
        "test_date": "17.08.2026",
        "report_date": "17.08.2026",
        "operator_name": "Ahmet Yilmaz",
        "operator_title": "Elektrik Muhendisi"
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(t3_payload, f, indent=2)
        
    cmd = [
        "python", os.path.join(TOOLS_DIR, "generate_excel.py"),
        "--json", json_path,
        "--template-type", "hermetik",
        "--has-breaker", "false",
        "--output", out_path
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    print(f"CLI Result: exit_code={res.returncode}")
    if res.returncode != 0:
        print(f"CLI stderr: {res.stderr}")
        return False
        
    wb = openpyxl.load_workbook(out_path, data_only=False)
    sheet_count = len(wb.sheetnames)
    has_kesici_sheet = any("KESİCİ" in s.upper() or "AÇMA" in s.upper() for s in wb.sheetnames)
    
    print(f"[T3] Total sheets: {sheet_count} (expected 12)")
    print(f"[T3] Breaker sheet present: {has_kesici_sheet} (expected False)")
    
    t3_success = (sheet_count == 12) and (not has_kesici_sheet)
    print(f"[T3] PASS: {t3_success}")
    return t3_success

def main():
    print("==================================================")
    print("EXECUTING ACCEPTANCE TESTS T1, T2, T3")
    print("==================================================")
    
    t1_ok = run_t1()
    t2_ok = run_t2()
    t3_ok = run_t3()
    
    print("\n==================================================")
    print(f"ALL TESTS PASSED: {t1_ok and t2_ok and t3_ok}")
    print("==================================================")

if __name__ == "__main__":
    main()
