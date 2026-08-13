import openpyxl

print('==================================================')
print('VERIFYING HYBRID EXCEL OUTPUTS & PRUNING')
print('==================================================')

# 1. Hermetik WITH breaker
wb_with = openpyxl.load_workbook('tools/test_out_hermetik_hybrid_with_breaker.xlsx', data_only=False)
print('Hermetik WITH Breaker Sheets:', wb_with.sheetnames)
has_b_sheets_with = any('KESİCİ' in s for s in wb_with.sheetnames)
print('  -> Has breaker sheets?', has_b_sheets_with)

# 2. Hermetik NO breaker
wb_no = openpyxl.load_workbook('tools/test_out_hermetik_hybrid_no_breaker.xlsx', data_only=False)
print('\nHermetik NO Breaker Sheets:', wb_no.sheetnames)
has_b_sheets_no = any('KESİCİ' in s for s in wb_no.sheetnames)
print('  -> Has breaker sheets?', has_b_sheets_no)

# 3. Kuru Tip Questions Verification
wb_kuru = openpyxl.load_workbook('tools/test_out_kuru_hybrid.xlsx', data_only=False)
ws_kuru_ana = wb_kuru['ANA SAYFA']
print('\nKuru Tip ANA SAYFA Checklist Sample Rows:')
for r in [26, 28, 30, 32, 34, 36]:
    print(f'  Row {r}: B{r}={repr(ws_kuru_ana[f"B{r}"].value)} | K{r}={repr(ws_kuru_ana[f"K{r}"].value)}')

# 4. Cover Page & Main Page Cell Values Inspection (Hermetik)
ws_kapak = wb_with['KAPAK SAYFASI']
print('\nHermetik KAPAK SAYFASI Cells:')
print('  D9 (customer_name):', repr(ws_kapak['D9'].value))
print('  D10 (trafo_label):', repr(ws_kapak['D10'].value))
print('  A29 (summary_text):', repr(ws_kapak['A29'].value))
print('  D55 (operator_title):', repr(ws_kapak['D55'].value))
print('  D56 (sicil_no):', repr(ws_kapak['D56'].value))
print('  D57 (ekipnet_no):', repr(ws_kapak['D57'].value))
print('  G52 (operator_name):', repr(ws_kapak['G52'].value))

ws_ana = wb_with['ANA SAYFA']
print('\nHermetik ANA SAYFA Cells:')
print('  F81 (operator_title):', repr(ws_ana['F81'].value))
print('  F82 (sicil_no):', repr(ws_ana['F82'].value))
print('  F83 (ekipnet_no):', repr(ws_ana['F83'].value))
print('  K78 (operator_name):', repr(ws_ana['K78'].value))
