import os
import shutil
import openpyxl
import copy
import io
from openpyxl.drawing.image import Image

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TEMPLATES_DIR = os.path.join(REPO_ROOT, "backend", "templates")
HYBRID_DIR = os.path.join(TEMPLATES_DIR, "hybrid")
MOBILE_TEMPLATES_DIR = os.path.join(REPO_ROOT, "mobile", "assets", "templates")

REF_PATH = os.path.join(TEMPLATES_DIR, "reference", "yeni_tasarim_referans.xlsx")
OLD_HERMETIK_PATH = os.path.join(TEMPLATES_DIR, "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx")
OLD_GT_PATH = os.path.join(TEMPLATES_DIR, "TR BAKIM RAPORU GT HİLMİ.xlsx")
OLD_KURU_PATH = os.path.join(TEMPLATES_DIR, "KURU TİP HİLMİ.xlsx")

SAMPLE_CLEAR_RANGES = {
    "ANA SAYFA KESİCİ": [
        "G11", "O11", "G13", "O13", "G15", "O15", "G17", "O17",
        "G24", "I24", "P24", "R24", "G26", "H26", "I26", "P26", "R26",
        "G28", "I28", "P28", "R28", "G30", "I30", "P30", "R30",
        "G32", "I32", "P32", "R32",
        "C39", "F39", "I39", "L39", "O39",
        "C41", "F41", "I41", "L41", "O41",
        "C43", "F43", "I43", "L43", "O43",
        "C50", "G50", "K50", "O50",
        "C52", "G52", "K52", "O52",
        "C54", "G54", "K54", "O54",
        "C61", "J61", "Q61", "C63", "J63", "Q63",
        "C65", "F65", "J65", "M65", "Q65",
        "F77", "F78",
    ],
    "OG SARGI MEVCUT KADEME": [
        "D11", "J11", "O11",
        "C24", "E24", "G24", "I24", "K24", "M24", "O24", "Q24",
        "C25", "E25", "G25", "I25", "K25", "M25", "O25", "Q25",
        "C26", "E26", "G26", "I26", "K26", "M26", "O26", "Q26",
    ],
    "AG SARGI": [
        "D11", "J11", "O11",
        "C24", "E24", "G24", "I24", "K24", "M24", "O24", "Q24",
        "C25", "E25", "G25", "I25", "K25", "M25", "O25", "Q25",
        "C26", "E26", "G26", "I26", "K26", "M26", "O26", "Q26",
    ],
    "İZOLASYON ": [
        "D11", "J11", "O11", "D16", "D17", "D30", "D31",
    ],
    "Ç.O 34500": [
        "D11", "J11", "O11",
        "B16", "C16", "D16", "B17", "C17", "D17",
        "B18", "C18", "D18", "B19", "C19", "D19", "B20", "C20", "D20",
    ],
    "TOPRAKLAMALAR": [
        "D9", "J9", "O9", "D17", "D18", "D19", "D32", "D33", "D34",
    ],
    "HV PF": ["D11", "J11", "O11", "P17"],
    "LV PF": ["D11", "J11", "O11", "P17"],
    "KESİCİ İZOLASYON": ["D10", "J10", "O10"],
    "KESİCİ KONTAK": ["D10", "J10", "O10"],
    "AÇMA-KAPAMA": ["D9", "J9", "O9", "D10"],
    "DİĞER": ["D9", "J9", "O9", "D16", "H16", "D17", "H17", "D18", "H18", "D23", "H23", "D24", "H24", "D25", "H25", "D30", "H30", "D31", "H31", "D32", "H32"],
    "AKIM TRAFOLARI": ["D9", "J9", "O9", "D16", "H16", "D17", "H17", "D18", "H18"],
}


def copy_sheet_content(src_ws, target_ws):
    """
    Copies cell values, styles, merged cells, row heights, column widths,
    sheet format, page setup, margins, print properties, and images with 100% dimension preservation.
    Ensures ALL rows have explicit locked height (customHeight=True) so Excel never auto-collapses or squishes rows.
    """
    def_rh = getattr(src_ws.sheet_format, 'defaultRowHeight', None) or 15.0

    # 1. Sheet format & default dimensions
    if hasattr(src_ws, 'sheet_format'):
        target_ws.sheet_format.defaultRowHeight = def_rh
        if src_ws.sheet_format.defaultColWidth is not None:
            target_ws.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth
        if src_ws.sheet_format.baseColWidth is not None:
            target_ws.sheet_format.baseColWidth = src_ws.sheet_format.baseColWidth

    # 2. Row dimensions for EVERY row from 1 to src_ws.max_row
    max_r = max(src_ws.max_row, max((r for r in src_ws.row_dimensions.keys()), default=1))
    for r in range(1, max_r + 1):
        rdim = src_ws.row_dimensions.get(r)
        trdim = target_ws.row_dimensions[r]
        if rdim and rdim.height is not None:
            trdim.height = rdim.height
        else:
            trdim.height = def_rh
        
        if rdim:
            if rdim.hidden:
                trdim.hidden = rdim.hidden
            if rdim.outline_level:
                trdim.outline_level = rdim.outline_level

    # 3. Column dimensions (width, hidden, outline_level)
    for c, cdim in src_ws.column_dimensions.items():
        tcdim = target_ws.column_dimensions[c]
        if cdim.width is not None:
            tcdim.width = cdim.width
        if cdim.hidden:
            tcdim.hidden = cdim.hidden
        if cdim.outline_level:
            tcdim.outline_level = cdim.outline_level

    # 4. Page Setup, Sheet View & Margins
    # Copy existing sheet_view properties if present
    if hasattr(src_ws, 'sheet_view') and src_ws.sheet_view is not None:
        sv_src = src_ws.sheet_view
        sv_tgt = target_ws.sheet_view
        if getattr(sv_src, 'view', None) is not None:
            sv_tgt.view = sv_src.view
        if getattr(sv_src, 'zoomScale', None) is not None:
            sv_tgt.zoomScale = sv_src.zoomScale
        if getattr(sv_src, 'zoomScaleNormal', None) is not None:
            sv_tgt.zoomScaleNormal = sv_src.zoomScaleNormal
        if getattr(sv_src, 'showGridLines', None) is not None:
            sv_tgt.showGridLines = sv_src.showGridLines

    # Page Setup
    if hasattr(src_ws, 'page_setup') and src_ws.page_setup is not None:
        for attr in ['orientation', 'paperSize', 'scale', 'fitToWidth', 'fitToHeight', 'firstPageNumber', 'useFirstPageNumber', 'pageOrder']:
            val = getattr(src_ws.page_setup, attr, None)
            if val is not None:
                setattr(target_ws.page_setup, attr, val)

    if hasattr(src_ws, 'sheet_properties') and hasattr(src_ws.sheet_properties, 'pageSetUpPr'):
        if src_ws.sheet_properties.pageSetUpPr.fitToPage is not None:
            target_ws.sheet_properties.pageSetUpPr.fitToPage = src_ws.sheet_properties.pageSetUpPr.fitToPage

    # Page Margins
    if hasattr(src_ws, 'page_margins') and src_ws.page_margins is not None:
        for attr in ['top', 'bottom', 'left', 'right', 'header', 'footer']:
            val = getattr(src_ws.page_margins, attr, None)
            if val is not None:
                setattr(target_ws.page_margins, attr, val)

    if hasattr(src_ws, 'print_title_rows') and src_ws.print_title_rows:
        target_ws.print_title_rows = src_ws.print_title_rows
    if hasattr(src_ws, 'print_title_cols') and src_ws.print_title_cols:
        target_ws.print_title_cols = src_ws.print_title_cols
    if hasattr(src_ws, 'print_area') and src_ws.print_area:
        target_ws.print_area = src_ws.print_area
    if hasattr(src_ws, 'freeze_panes') and src_ws.freeze_panes:
        target_ws.freeze_panes = src_ws.freeze_panes

    # 5. Cell values & styles
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, min_col=1, max_col=src_ws.max_column):
        for cell in row:
            tcell = target_ws.cell(row=cell.row, column=cell.column)
            tcell.value = cell.value
            if cell.has_style:
                tcell.font = copy.copy(cell.font)
                tcell.border = copy.copy(cell.border)
                tcell.fill = copy.copy(cell.fill)
                tcell.number_format = cell.number_format
                tcell.protection = copy.copy(cell.protection)
                tcell.alignment = copy.copy(cell.alignment)

    # 6. Merged Cells
    for rng in src_ws.merged_cells.ranges:
        try:
            target_ws.merge_cells(str(rng))
        except Exception:
            pass

    # 7. Images
    if hasattr(src_ws, '_images') and src_ws._images:
        for img in src_ws._images:
            try:
                r = getattr(img.anchor._from, 'row', None) if hasattr(img, 'anchor') and hasattr(img.anchor, '_from') else None
                c = getattr(img.anchor._from, 'col', None) if hasattr(img, 'anchor') and hasattr(img.anchor, '_from') else None
                if src_ws.title == "KAPAK SAYFASI" and r is not None and r > 2:
                    continue
                if r is not None and c is not None and r >= 30 and c >= 5:
                    continue
                img_data = img._data()
                new_img = Image(io.BytesIO(img_data))
                if hasattr(img, 'anchor'):
                    new_img.anchor = copy.copy(img.anchor)
                target_ws.add_image(new_img)
            except Exception as e:
                print(f"Warning copying image in sheet {src_ws.title}: {e}")


def clean_master_images(wb):
    """Removes old field sample photos and sample signature drawings from master templates."""
    for sname in wb.sheetnames:
        ws = wb[sname]
        if hasattr(ws, '_images') and ws._images:
            clean = []
            for img in ws._images:
                r = getattr(img.anchor._from, 'row', None) if hasattr(img, 'anchor') and hasattr(img.anchor, '_from') else None
                c = getattr(img.anchor._from, 'col', None) if hasattr(img, 'anchor') and hasattr(img.anchor, '_from') else None
                if sname == "KAPAK SAYFASI":
                    if r is not None and r <= 2:
                        clean.append(img)
                else:
                    is_old_sig = False
                    if r is not None and c is not None:
                        if r >= 30 and c >= 5:
                            is_old_sig = True
                    if not is_old_sig:
                        clean.append(img)
            ws._images = clean


def clean_master_sample_data(wb):
    """Clears hardcoded sample data cells from master hybrid template sheets at build time."""
    for sname, cell_list in SAMPLE_CLEAR_RANGES.items():
        matched_sheet = None
        for name in wb.sheetnames:
            if name.strip() == sname.strip():
                matched_sheet = name
                break
        if matched_sheet:
            ws = wb[matched_sheet]
            for cref in cell_list:
                try:
                    cell = ws[cref]
                    if cell.value is not None and not str(cell.value).strip().startswith('='):
                        cell.value = None
                except Exception:
                    pass


def fix_xlsx_rels(xlsx_path: str):
    """
    Fixes relationship target paths in .xlsx zip archive by converting
    absolute targets starting with '/xl/' (e.g. '/xl/drawings/drawing1.xml', '/xl/media/image1.png')
    to standard relative targets without leading '/xl/'.
    """
    try:
        import zipfile
        buffer = io.BytesIO()
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    content = zin.read(item.filename)
                    if item.filename.endswith('.rels'):
                        try:
                            import xml.etree.ElementTree as ET
                            root = ET.fromstring(content)
                            modified = False
                            parts = item.filename.split('/')
                            if len(parts) >= 2 and parts[-2] == '_rels':
                                base_dir = '/'.join(parts[:-2])
                            else:
                                base_dir = '/'.join(parts[:-1])

                            for child in root:
                                tgt = child.get('Target')
                                if tgt and tgt.startswith('/xl/'):
                                    full_target = tgt[1:]  # strip leading '/'
                                    rel_target = os.path.relpath(full_target, base_dir).replace('\\', '/')
                                    child.set('Target', rel_target)
                                    modified = True

                            if modified:
                                content = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                        except Exception:
                            pass
                    zout.writestr(item, content)

        with open(xlsx_path, 'wb') as f:
            f.write(buffer.getvalue())
    except Exception as e:
        print(f"Warning fixing rels in {xlsx_path}: {e}")


def sync_to_all_targets(src_hybrid, old_template, mobile_name):
    """Copies the newly built hybrid template to legacy paths and mobile assets."""
    shutil.copyfile(src_hybrid, old_template)
    fix_xlsx_rels(old_template)
    
    if os.path.exists(MOBILE_TEMPLATES_DIR):
        mob1 = os.path.join(MOBILE_TEMPLATES_DIR, mobile_name)
        mob2 = os.path.join(MOBILE_TEMPLATES_DIR, os.path.basename(src_hybrid))
        shutil.copyfile(src_hybrid, mob1)
        shutil.copyfile(src_hybrid, mob2)
        fix_xlsx_rels(mob1)
        fix_xlsx_rels(mob2)
        print(f"Synced {os.path.basename(src_hybrid)} -> {mob1} and {mob2}")


def build_hermetik_hybrid():
    print("Building hermetik_hybrid.xlsx...")
    wb_ref = openpyxl.load_workbook(REF_PATH, data_only=False)
    wb_old = openpyxl.load_workbook(OLD_HERMETIK_PATH, data_only=False)

    for sname in wb_old.sheetnames:
        if sname not in wb_ref.sheetnames:
            target_ws = wb_ref.create_sheet(title=sname)
            copy_sheet_content(wb_old[sname], target_ws)

    original_order = wb_old.sheetnames
    ordered_sheets = []
    for sname in original_order:
        if sname in wb_ref.sheetnames:
            ordered_sheets.append(wb_ref[sname])
    for sname in wb_ref.sheetnames:
        if wb_ref[sname] not in ordered_sheets:
            ordered_sheets.append(wb_ref[sname])
    wb_ref._sheets = ordered_sheets

    clean_master_images(wb_ref)
    clean_master_sample_data(wb_ref)

    out_hybrid = os.path.join(HYBRID_DIR, "hermetik_hybrid.xlsx")
    wb_ref.save(out_hybrid)
    fix_xlsx_rels(out_hybrid)
    sync_to_all_targets(out_hybrid, OLD_HERMETIK_PATH, "hermetik.xlsx")
    print(f"Saved: {out_hybrid} and updated all targets.")


def build_gt_hybrid():
    print("Building gt_hybrid.xlsx...")
    wb_ref = openpyxl.load_workbook(REF_PATH, data_only=False)
    wb_old = openpyxl.load_workbook(OLD_GT_PATH, data_only=False)

    for sname in wb_old.sheetnames:
        if sname not in wb_ref.sheetnames:
            target_ws = wb_ref.create_sheet(title=sname)
            copy_sheet_content(wb_old[sname], target_ws)

    original_order = wb_old.sheetnames
    ordered_sheets = []
    for sname in original_order:
        if sname in wb_ref.sheetnames:
            ordered_sheets.append(wb_ref[sname])
    for sname in wb_ref.sheetnames:
        if wb_ref[sname] not in ordered_sheets:
            ordered_sheets.append(wb_ref[sname])
    wb_ref._sheets = ordered_sheets

    clean_master_images(wb_ref)
    clean_master_sample_data(wb_ref)

    out_hybrid = os.path.join(HYBRID_DIR, "gt_hybrid.xlsx")
    wb_ref.save(out_hybrid)
    fix_xlsx_rels(out_hybrid)
    sync_to_all_targets(out_hybrid, OLD_GT_PATH, "gt.xlsx")
    print(f"Saved: {out_hybrid} and updated all targets.")


def build_kuru_tip_hybrid():
    print("Building kuru_tip_hybrid.xlsx...")
    wb_ref = openpyxl.load_workbook(REF_PATH, data_only=False)
    wb_kuru = openpyxl.load_workbook(OLD_KURU_PATH, data_only=False)

    for sname in wb_kuru.sheetnames:
        if sname not in ["KAPAK SAYFASI"] and sname not in wb_ref.sheetnames:
            target_ws = wb_ref.create_sheet(title=sname)
            copy_sheet_content(wb_kuru[sname], target_ws)

    clean_master_images(wb_ref)
    clean_master_sample_data(wb_ref)

    out_hybrid = os.path.join(HYBRID_DIR, "kuru_tip_hybrid.xlsx")
    wb_ref.save(out_hybrid)
    fix_xlsx_rels(out_hybrid)
    sync_to_all_targets(out_hybrid, OLD_KURU_PATH, "kuru_tip.xlsx")
    print(f"Saved: {out_hybrid} and updated all targets.")


def main():
    os.makedirs(HYBRID_DIR, exist_ok=True)
    os.makedirs(MOBILE_TEMPLATES_DIR, exist_ok=True)
    build_hermetik_hybrid()
    build_gt_hybrid()
    build_kuru_tip_hybrid()
    print("Master hybrid template generation completed successfully.")


if __name__ == "__main__":
    main()
