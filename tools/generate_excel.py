#!/usr/bin/env python3
"""
tools/generate_excel.py
Standalone Python CLI tool for generating professional Excel transformer report (.xlsx) files
using openpyxl and official template files.
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from typing import Dict, Optional, Any

try:
    import openpyxl
    from openpyxl.drawing.image import Image as OpenPyXLImage
    from openpyxl.styles import Font, Alignment
except ImportError:
    sys.stderr.write("ERROR: openpyxl is not installed. Run 'pip install openpyxl'\n")
    sys.exit(1)

# Type-specific cell mappings per sheet matching docs/EXCEL_CELL_MAPPING.md and excel_engine.py
TYPE_CELL_MAPPINGS: Dict[str, Dict[str, Dict[str, str]]] = {
    "HERMETIK": {
        "KAPAK SAYFASI": {
            "D9": "customer_name",
            "D10": "trafo_label",
            "D11": "address",
            "D12": "report_date",
            "D14": "test_date",
            "A29": "summary_text",
            "D55": "operator_title",
        },
        "ANA SAYFA": {
            "G11": "brand",
            "O11": "tap_info_1",
            "Q11": "tap_info_2",
            "S11": "tap_info_3",
            "G13": "power_kva",
            "O13": "manufacture_year",
            "G15": "voltage",
            "O15": "serial_no",
            "G17": "oil_brand",
            "G31": "dc_redresor_voltage",
            "O17": "oil_weight",
            "G19": "connection_group",
            "O19": "short_circuit_imp_pct",
            "G21": "tank_type",
            "I21": "tank_mark_hermetik",
            "P21": "tank_mark_gt",
            "U21": "tank_mark_kuru",
            "C55": "og_rab",
            "C57": "og_rbc",
            "C59": "og_rca",
            "J55": "ag_ran",
            "J57": "ag_rbn",
            "J59": "ag_rcn",
            "O55": "ag_rab",
            "O57": "ag_rbc",
            "O59": "ag_rca",
            "F66": "iso_yg_tank_30s",
            "I66": "iso_yg_tank_45s",
            "L66": "iso_yg_tank_60s",
            "F68": "iso_ag_tank_30s",
            "I68": "iso_ag_tank_45s",
            "L68": "iso_ag_tank_60s",
            "F70": "iso_yg_ag_30s",
            "I70": "iso_yg_ag_45s",
            "L70": "iso_yg_ag_60s",
            "H46": "ground_koruma_trafo",
            "H48": "ground_isletme_notr",
            "H50": "ground_koruma_hucre",
            "Q46": "ground_koruma_ag_pano",
            "Q48": "ground_koruma_kapilar",
            "B73": "notes",
            "F81": "operator_title",
        },
        "OG SARGI MEVCUT KADEME": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "AG SARGI": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "İZOLASYON ": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "D16": "iso_og_gnd",
            "D17": "iso_ag_gnd",
            "D30": "iso_temp",
            "D31": "iso_humidity",
        },
        "Ç.O 34500": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "B16": "ttr_tap1_a",
            "C16": "ttr_tap1_b",
            "D16": "ttr_tap1_c",
            "B17": "ttr_tap2_a",
            "C17": "ttr_tap2_b",
            "D17": "ttr_tap2_c",
            "B18": "ttr_tap3_a",
            "C18": "ttr_tap3_b",
            "D18": "ttr_tap3_c",
            "B19": "ttr_tap4_a",
            "C19": "ttr_tap4_b",
            "D19": "ttr_tap4_c",
            "B20": "ttr_tap5_a",
            "C20": "ttr_tap5_b",
            "D20": "ttr_tap5_c",
        },
        "TOPRAKLAMALAR": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "K16": "ground_isletme_notr",
            "K26": "ground_koruma_trafo",
            "K27": "ground_koruma_hucre",
            "K28": "ground_koruma_kapilar",
            "K29": "ground_koruma_ag_pano",
        },
        "HV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_hv_humidity",
        },
        "LV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_lv_humidity",
        },
        "ANA SAYFA KESİCİ": {
            "G11": "breaker_brand",
            "O11": "breaker_serial_no",
            "G13": "breaker_model",
            "O13": "breaker_year",
            "G15": "breaker_voltage",
            "G17": "breaker_motor_voltage",
            "O17": "breaker_coil_voltage",
            "G26": "breaker_dc_redresor_voltage",
            "G50": "breaker_iso_r_gnd",
            "O50": "breaker_contact_r",
            "D68": "breaker_notes",
            "F76": "operator_title",
        },
        "KESİCİ İZOLASYON": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "KESİCİ KONTAK": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "AÇMA-KAPAMA": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D10": "breaker_timing_open",
        },
        "DİĞER": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
        },
        "AKIM TRAFOLARI": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D16": "ct_ratio",
        },
        "HERMETİK YAĞ DİLEKÇESİ": {}
    },
    "KURU_TIP": {
        "KAPAK SAYFASI": {
            "D9": "customer_name",
            "D10": "trafo_label",
            "D11": "address",
            "D12": "report_date",
            "D14": "test_date",
            "A29": "summary_text",
            "D55": "operator_title",
        },
        "ANA SAYFA": {
            "G11": "brand",
            "O11": "tap_info_1",
            "Q11": "tap_info_2",
            "S11": "tap_info_3",
            "G13": "power_kva",
            "O13": "manufacture_year",
            "G15": "voltage",
            "O15": "serial_no",
            "G17": "connection_group",
            "G31": "dc_redresor_voltage",
            "O17": "short_circuit_imp_pct",
            "G19": "tank_type",
            "I19": "tank_mark_hermetik",
            "P19": "tank_mark_gt",
            "U19": "tank_mark_kuru",
            "C49": "og_rab",
            "C51": "og_rbc",
            "C53": "og_rca",
            "J49": "ag_ran",
            "J51": "ag_rbn",
            "J53": "ag_rcn",
            "O49": "ag_rab",
            "O51": "ag_rbc",
            "O53": "ag_rca",
            "B67": "notes",
            "F74": "operator_title",
        },
        "OG SARGI MEVCUT KADEME": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "AG SARGI": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "İZOLASYON ": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "D16": "iso_og_gnd",
            "D17": "iso_ag_gnd",
            "D30": "iso_temp",
            "D31": "iso_humidity",
        },
        "Ç.O 34500": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "B16": "ttr_tap1_a",
            "C16": "ttr_tap1_b",
            "D16": "ttr_tap1_c",
            "B17": "ttr_tap2_a",
            "C17": "ttr_tap2_b",
            "D17": "ttr_tap2_c",
            "B18": "ttr_tap3_a",
            "C18": "ttr_tap3_b",
            "D18": "ttr_tap3_c",
            "B19": "ttr_tap4_a",
            "C19": "ttr_tap4_b",
            "D19": "ttr_tap4_c",
            "B20": "ttr_tap5_a",
            "C20": "ttr_tap5_b",
            "D20": "ttr_tap5_c",
        },
        "TOPRAKLAMALAR": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "K16": "ground_isletme_notr",
            "K26": "ground_koruma_trafo",
            "K27": "ground_koruma_hucre",
            "K28": "ground_koruma_kapilar",
            "K29": "ground_koruma_ag_pano",
        },
        "HV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_hv_humidity",
        },
        "LV PF": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "P17": "pf_lv_humidity",
        },
        "ANA SAYFA KESİCİ": {
            "G11": "breaker_brand",
            "O11": "breaker_serial_no",
            "G13": "breaker_model",
            "O13": "breaker_year",
            "G15": "breaker_voltage",
            "G17": "breaker_motor_voltage",
            "O17": "breaker_coil_voltage",
            "G26": "breaker_dc_redresor_voltage",
            "G50": "breaker_iso_r_gnd",
            "O50": "breaker_contact_r",
            "D68": "breaker_notes",
            "F76": "operator_title",
        },
        "KESİCİ İZOLASYON": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "KESİCİ KONTAK": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "AÇMA-KAPAMA": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D10": "breaker_timing_open",
        },
        "DİĞER": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
        },
        "AKIM TRAFOLARI": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D16": "ct_ratio",
        }
    },
    "GT": {
        "KAPAK SAYFASI": {
            "D9": "customer_name",
            "D10": "trafo_label",
            "D11": "address",
            "D12": "report_date",
            "D14": "test_date",
            "A29": "summary_text",
            "D55": "operator_title",
        },
        "ANA SAYFA": {
            "G11": "brand",
            "O11": "tap_info_1",
            "Q11": "tap_info_2",
            "S11": "tap_info_3",
            "G13": "power_kva",
            "O13": "manufacture_year",
            "G15": "voltage",
            "O15": "serial_no",
            "G17": "oil_brand",
            "G31": "dc_redresor_voltage",
            "O17": "oil_weight",
            "G19": "connection_group",
            "O19": "short_circuit_imp_pct",
            "G21": "tank_type",
            "I21": "tank_mark_hermetik",
            "P21": "tank_mark_gt",
            "U21": "tank_mark_kuru",
            "C55": "og_rab",
            "C57": "og_rbc",
            "C59": "og_rca",
            "J55": "ag_ran",
            "J57": "ag_rbn",
            "J59": "ag_rcn",
            "O55": "ag_rab",
            "O57": "ag_rbc",
            "O59": "ag_rca",
            "F66": "iso_yg_tank_30s",
            "I66": "iso_yg_tank_45s",
            "L66": "iso_yg_tank_60s",
            "F68": "iso_ag_tank_30s",
            "I68": "iso_ag_tank_45s",
            "L68": "iso_ag_tank_60s",
            "F70": "iso_yg_ag_30s",
            "I70": "iso_yg_ag_45s",
            "L70": "iso_yg_ag_60s",
            "H46": "ground_koruma_trafo",
            "H48": "ground_isletme_notr",
            "H50": "ground_koruma_hucre",
            "Q46": "ground_koruma_ag_pano",
            "Q48": "ground_koruma_kapilar",
            "B73": "notes",
            "F81": "operator_title",
        },
        "OG SARGI MEVCUT KADEME": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "AG SARGI": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
        },
        "İZOLASYON ": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "D16": "iso_og_gnd",
            "D17": "iso_ag_gnd",
            "D30": "iso_temp",
            "D31": "iso_humidity",
        },
        "Ç.O 34500": {
            "D11": "operator_name",
            "J11": "device_model",
            "O11": "device_serial",
            "B16": "ttr_tap1_a",
            "C16": "ttr_tap1_b",
            "D16": "ttr_tap1_c",
            "B17": "ttr_tap2_a",
            "C17": "ttr_tap2_b",
            "D17": "ttr_tap2_c",
            "B18": "ttr_tap3_a",
            "C18": "ttr_tap3_b",
            "D18": "ttr_tap3_c",
            "B19": "ttr_tap4_a",
            "C19": "ttr_tap4_b",
            "D19": "ttr_tap4_c",
            "B20": "ttr_tap5_a",
            "C20": "ttr_tap5_b",
            "D20": "ttr_tap5_c",
        },
        "TOPRAKLAMALAR": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "K16": "ground_isletme_notr",
            "K26": "ground_koruma_trafo",
            "K27": "ground_koruma_hucre",
            "K28": "ground_koruma_kapilar",
            "K29": "ground_koruma_ag_pano",
        },
        "ANA SAYFA KESİCİ": {
            "G11": "breaker_brand",
            "O11": "breaker_serial_no",
            "G13": "breaker_model",
            "O13": "breaker_year",
            "G15": "breaker_voltage",
            "G17": "breaker_motor_voltage",
            "O17": "breaker_coil_voltage",
            "G26": "breaker_dc_redresor_voltage",
            "G50": "breaker_iso_r_gnd",
            "O50": "breaker_contact_r",
            "D68": "breaker_notes",
            "F76": "operator_title",
        },
        "KESİCİ İZOLASYON": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "KESİCİ KONTAK": {
            "D10": "operator_name",
            "J10": "device_model",
            "O10": "device_serial",
        },
        "AÇMA-KAPAMA": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D10": "breaker_timing_open",
        },
        "DİĞER": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
        },
        "AKIM TRAFOLARI": {
            "D9": "operator_name",
            "J9": "device_model",
            "O9": "device_serial",
            "D16": "ct_ratio",
        },
        "YAĞ RAPORU": {
            "D16": "oil_test_breakdown_voltage",
            "D18": "oil_test_water_content",
        },
        "HERMETİK YAĞ DİLEKÇESİ": {}
    }
}

SAMPLE_CLEAR_RANGES = {
    "TOPRAKLAMALAR": ["K16", "K26", "K27", "K28", "K29"],
    "ANA SAYFA": [
        "G31", "O55", "O57", "O59",
        "F66", "I66", "L66",
        "F68", "I68", "L68",
        "F70", "I70", "L70",
        "H46", "H48", "H50", "Q46", "Q48",
        "H40", "H42", "H44", "Q40", "Q42"
    ],
    "ANA SAYFA KESİCİ": [
        "G11", "O11", "G13", "O13", "G15", "O15", "G17", "O17",
        "G24", "I24", "P24", "R24", "G26", "H26", "I26", "P26", "R26",
        "G28", "I28", "P28", "R28", "G30", "I30", "P30", "R30",
        "G32", "I32", "P32", "R32",
        "C39", "F39", "I39", "L39", "O39",
        "C41", "F41", "I41", "L41", "O41",
        "C43", "F43", "I43", "L43", "O43",
        "C50", "G50", "K50", "O50",
        "C52", "G52", "K52", "O52",
        "C54", "G54", "K54", "O54",
        "C61", "J61", "Q61", "C63", "J63", "Q63",
        "C65", "F65", "J65", "M65", "Q65",
        "F77", "F78",
    ],
    "OG SARGI MEVCUT KADEME": [
        "D11", "J11", "O11",
        "C24", "E24", "G24", "I24", "K24", "M24", "O24", "Q24",
        "C25", "E25", "G25", "I25", "K25", "M25", "O25", "Q25",
        "C26", "E26", "G26", "I26", "K26", "M26", "O26", "Q26",
    ],
    "AG SARGI": [
        "D11", "J11", "O11",
        "C24", "E24", "G24", "I24", "K24", "M24", "O24", "Q24",
        "C25", "E25", "G25", "I25", "K25", "M25", "O25", "Q25",
        "C26", "E26", "G26", "I26", "K26", "M26", "O26", "Q26",
    ],
    "İZOLASYON ": [
        "D11", "J11", "O11", "D16", "D17", "D30", "D31",
    ],
    "Ç.O 34500": [
        "D11", "J11", "O11",
        "B16", "C16", "D16", "B17", "C17", "D17",
        "B18", "C18", "D18", "B19", "C19", "D19", "B20", "C20", "D20",
    ],
    "TOPRAKLAMALAR": [
        "D9", "J9", "O9", "D17", "D18", "D19", "D32", "D33", "D34",
    ],
    "HV PF": ["D11", "J11", "O11", "P17"],
    "LV PF": ["D11", "J11", "O11", "P17"],
    "KESİCİ İZOLASYON": ["D10", "J10", "O10"],
    "KESİCİ KONTAK": ["D10", "J10", "O10"],
    "AÇMA-KAPAMA": ["D9", "J9", "O9", "D10"],
    "DİĞER": ["D9", "J9", "O9", "D16", "H16", "D17", "H17", "D18", "H18", "D23", "H23", "D24", "H24", "D25", "H25", "D30", "H30", "D31", "H31", "D32", "H32"],
    "AKIM TRAFOLARI": ["D9", "J9", "O9", "D16", "H16", "D17", "H17", "D18", "H18"],
}


def pre_clear_sample_cells(wb):
    """
    Pre-clears hardcoded sample data cells across all sheets in wb.
    Preserves all formulas, titles, labels, logos, formatting, and row heights.
    """
    for sname, cell_list in SAMPLE_CLEAR_RANGES.items():
        matched_sheet = None
        for name in wb.sheetnames:
            if name.strip() == sname.strip():
                matched_sheet = name
                break
        if matched_sheet:
            ws = wb[matched_sheet]
            for cref in cell_list:
                try:
                    cell = ws[cref]
                    if type(cell).__name__ != 'MergedCell' and cell.value is not None:
                        if not str(cell.value).strip().startswith('='):
                            cell.value = None
                except Exception:
                    pass


# Sheet signature anchors disabled across all sheets (Rule C)
SHEET_SIGNATURE_ANCHORS = {
    "HERMETIK": {},
    "KURU_TIP": {},
    "GT": {}
}

CHECKLIST_PAIRS = {
    "HERMETIK": {
        "checklist_1": {"evet": "I27", "hayir": "J27"},
        "checklist_2": {"evet": "I29", "hayir": "J29"},
        "checklist_3": {"evet": "I31", "hayir": "J31"},
        "checklist_4": {"evet": "I33", "hayir": "J33"},
        "checklist_5": {"evet": "I35", "hayir": "J35"},
        "checklist_6": {"evet": "I37", "hayir": "J37"},
        "checklist_7": {"evet": "I39", "hayir": "J39"},
        "checklist_8": {"evet": "I41", "hayir": "J41"},
        "checklist_9": {"evet": "R27", "hayir": "S27"},
        "checklist_10": {"evet": "R29", "hayir": "S29"},
        "checklist_11": {"evet": "R31", "hayir": "S31"},
        "checklist_12": {"evet": "R33", "hayir": "S33"},
        "checklist_13": {"evet": "R35", "hayir": "S35"},
        "checklist_14": {"evet": "R37", "hayir": "S37"},
        "checklist_15": {"evet": "R39", "hayir": "S39"},
        "checklist_16": {"evet": "R41", "hayir": "S41"},
        "breaker_control_visual": {"evet": "G24", "hayir": "I24"},
        "breaker_control_cleanliness": {"evet": "P24", "hayir": "R24"},
        "breaker_control_dc_redresor": {"evet": "I26", "hayir": "J26"},
        "breaker_control_cell_cleanliness": {"evet": "P26", "hayir": "R26"},
        "breaker_control_indicator": {"evet": "G28", "hayir": "I28"},
        "breaker_control_busbar": {"evet": "P28", "hayir": "R28"},
        "breaker_control_mechanical": {"evet": "G30", "hayir": "I30"},
        "breaker_control_heater": {"evet": "P30", "hayir": "R30"},
        "breaker_control_cable": {"evet": "G32", "hayir": "I32"},
        "breaker_control_relay": {"evet": "P32", "hayir": "R32"},
    },
    "KURU_TIP": {
        "checklist_1": {"evet": "I27", "hayir": "J27"},
        "checklist_2": {"evet": "I29", "hayir": "J29"},
        "checklist_3": {"evet": "I31", "hayir": "J31"},
        "checklist_4": {"evet": "I33", "hayir": "J33"},
        "checklist_5": {"evet": "I35", "hayir": "J35"},
        "checklist_6": {"evet": "I37", "hayir": "J37"},
        "checklist_7": {"evet": "R27", "hayir": "S27"},
        "checklist_8": {"evet": "R29", "hayir": "S29"},
        "checklist_9": {"evet": "R31", "hayir": "S31"},
        "checklist_10": {"evet": "R33", "hayir": "S33"},
        "checklist_11": {"evet": "R35", "hayir": "S35"},
        "checklist_12": {"evet": "R37", "hayir": "S37"},
        "breaker_control_visual": {"evet": "G24", "hayir": "I24"},
        "breaker_control_cleanliness": {"evet": "P24", "hayir": "R24"},
        "breaker_control_dc_redresor": {"evet": "I26", "hayir": "J26"},
        "breaker_control_cell_cleanliness": {"evet": "P26", "hayir": "R26"},
        "breaker_control_indicator": {"evet": "G28", "hayir": "I28"},
        "breaker_control_busbar": {"evet": "P28", "hayir": "R28"},
        "breaker_control_mechanical": {"evet": "G30", "hayir": "I30"},
        "breaker_control_heater": {"evet": "P30", "hayir": "R30"},
        "breaker_control_cable": {"evet": "G32", "hayir": "I32"},
        "breaker_control_relay": {"evet": "P32", "hayir": "R32"},
    },
    "GT": {
        "checklist_1": {"evet": "I27", "hayir": "J27"},
        "checklist_2": {"evet": "I29", "hayir": "J29"},
        "checklist_3": {"evet": "I31", "hayir": "J31"},
        "checklist_4": {"evet": "I33", "hayir": "J33"},
        "checklist_5": {"evet": "I35", "hayir": "J35"},
        "checklist_6": {"evet": "I37", "hayir": "J37"},
        "checklist_7": {"evet": "I39", "hayir": "J39"},
        "checklist_8": {"evet": "I41", "hayir": "J41"},
        "checklist_9": {"evet": "R27", "hayir": "S27"},
        "checklist_10": {"evet": "R29", "hayir": "S29"},
        "checklist_11": {"evet": "R31", "hayir": "S31"},
        "checklist_12": {"evet": "R33", "hayir": "S33"},
        "checklist_13": {"evet": "R35", "hayir": "S35"},
        "checklist_14": {"evet": "R37", "hayir": "S37"},
        "checklist_15": {"evet": "R39", "hayir": "S39"},
        "checklist_16": {"evet": "R41", "hayir": "S41"},
        "breaker_control_visual": {"evet": "G24", "hayir": "I24"},
        "breaker_control_cleanliness": {"evet": "P24", "hayir": "R24"},
        "breaker_control_dc_redresor": {"evet": "G26", "hayir": "I26"},
        "breaker_control_cell_cleanliness": {"evet": "P26", "hayir": "R26"},
        "breaker_control_indicator": {"evet": "G28", "hayir": "I28"},
        "breaker_control_busbar": {"evet": "P28", "hayir": "R28"},
        "breaker_control_mechanical": {"evet": "G30", "hayir": "I30"},
        "breaker_control_heater": {"evet": "P30", "hayir": "R30"},
        "breaker_control_cable": {"evet": "G32", "hayir": "I32"},
        "breaker_control_relay": {"evet": "P32", "hayir": "R32"},
    }
}


def date_to_excel_serial(date_input: Any) -> Optional[float]:
    """Converts DateTime or string (DD.MM.YYYY or YYYY-MM-DD) into Excel serial date number."""
    if date_input is None:
        return None
    try:
        if isinstance(date_input, (int, float)):
            return float(date_input)

        s_date = str(date_input).strip()
        if not s_date:
            return None

        dt_obj = None
        if "." in s_date:
            parts = s_date.split(".")
            if len(parts) == 3:
                p1, p2, p3 = int(parts[0]), int(parts[1]), int(parts[2])
                if p3 > 1000:
                    dt_obj = datetime(p3, p2, p1)
                elif p1 > 1000:
                    dt_obj = datetime(p1, p2, p3)
        elif "-" in s_date:
            parts = s_date.split("-")
            if len(parts) == 3:
                dt_obj = datetime(int(parts[0]), int(parts[1]), int(parts[2]))

        if dt_obj is not None:
            epoch = datetime(1899, 12, 30)
            return float((dt_obj - epoch).days)
    except Exception:
        pass
    return None


def get_writable_cell(ws, cell_ref: str):
    """Returns the top-left cell if cell_ref belongs to a merged range."""
    cell = ws[cell_ref]
    if type(cell).__name__ == 'MergedCell':
        for rng in ws.merged_cells.ranges:
            if cell_ref in rng:
                return ws.cell(rng.min_row, rng.min_col)
    return cell


def process_checklist_pairs(ws, report_type, data_dict):
    """
    Sets Evet/Hayır checkmarks for checklist items and clears residual template sample checkmarks.
    Also clears 3.0 pt divider row cells C61, F61, J61, C63, F63, J63 on ANA SAYFA to prevent ghost text.
    """
    for cref in ["C61", "F61", "J61", "C63", "F63", "J63"]:
        try:
            cell = get_writable_cell(ws, cref)
            cell.value = None
        except Exception:
            pass

    pairs_map = CHECKLIST_PAIRS.get(report_type, CHECKLIST_PAIRS["HERMETIK"])
    for key, pair in pairs_map.items():
        evet_cell = get_writable_cell(ws, pair["evet"])
        hayir_cell = get_writable_cell(ws, pair["hayir"])

        val = data_dict.get(key)
        if val is None:
            evet_cell.value = None
            hayir_cell.value = None
        else:
            is_true = (val is True or str(val).strip().lower() in ["true", "ü", "1", "evet"])
            is_false = (val is False or str(val).strip().lower() in ["false", "0", "hayir", "hayır"])
            if is_true:
                evet_cell.value = "ü"
                hayir_cell.value = None
            elif is_false:
                evet_cell.value = None
                hayir_cell.value = "ü"
            else:
                evet_cell.value = None
                hayir_cell.value = None


def process_sheet_signature(ws, target_anchor: Optional[str], sig_path: Optional[str]):
    """
    Cleans sample text artifacts from the worksheet.
    DOES NOT insert any user signature image (signature placement is disabled across all sheets).
    Clears all right block cells below 'ONAYLAYAN' header.
    """
    # Clean sample texts in right ONAYLAYAN block
    for r in range(40, min(86, ws.max_row + 1)):
        for c in range(5, 15):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell' and cell.value is not None:
                cval_str = str(cell.value).strip()
                cval_u = cval_str.upper()
                if "ONAYLAYAN" in cval_u:
                    # Keep cell.value as 'ONAYLAYAN', but clear body cells below it
                    for br in range(r + 1, min(r + 7, ws.max_row + 1)):
                        for bc in range(c, min(c + 7, 20)):
                            bcell = ws.cell(row=br, column=bc)
                            if type(bcell).__name__ != 'MergedCell':
                                bcell.value = None


def process_kapak_photos(ws_kapak, data_dict, args):
    """
    Inserts loaded photos (photo_before, photo_after, photo_label, photo_extra)
    into KAPAK SAYFASI and writes bold Turkish captions beneath each photo slot.
    """
    # Extract photos
    p_before = getattr(args, 'photo_before', None) or data_dict.get("photo_before")
    p_after = getattr(args, 'photo_after', None) or data_dict.get("photo_after")
    p_label = getattr(args, 'photo_label', None) or data_dict.get("photo_label")
    p_extra = data_dict.get("photo_extra")

    photos_to_place = []
    if p_before and isinstance(p_before, str) and os.path.exists(p_before):
        photos_to_place.append((p_before, "Bakım Öncesi"))
    if p_after and isinstance(p_after, str) and os.path.exists(p_after):
        photos_to_place.append((p_after, "Bakım Sonrası"))
    if p_label and isinstance(p_label, str) and os.path.exists(p_label):
        photos_to_place.append((p_label, "Trafo Etiket / Plaka"))
    if p_extra and isinstance(p_extra, str) and os.path.exists(p_extra):
        photos_to_place.append((p_extra, "Ek Fotoğraf"))

    # Slots grid: (anchor_cell, caption_cell, width, height)
    # Personnel block starts at row 52. All photo slots and captions stay strictly above row 52!
    # Slot 3 (photo_label) size equalized to 210x135 px like Slot 1 & 2!
    slots = [
        ("A32", "A38", 200, 120),
        ("G32", "G38", 200, 120),
        ("A40", "A47", 200, 115),
        ("G40", "G47", 200, 115),
    ]

    # Clear all potential caption cells first
    for cref in ["A38", "G38", "A41", "G41", "A47", "G47", "A50", "G50"]:
        try:
            cell = get_writable_cell(ws_kapak, cref)
            cell.value = None
        except Exception:
            pass

    font_bold = Font(name="Calibri", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Merge caption ranges to prevent any text truncation (e.g. 'Trafo Etiket / Plaka')
    caption_merges = [("A38", "A38:D38"), ("G38", "G38:J38"), ("A47", "A47:D47"), ("G47", "G47:J47")]
    for cref, mrange in caption_merges:
        try:
            ws_kapak.merge_cells(mrange)
        except Exception:
            pass

    for idx, (p_path, label_text) in enumerate(photos_to_place):
        if idx < len(slots):
            anchor_cell, caption_ref, w_px, h_px = slots[idx]
            try:
                img = OpenPyXLImage(p_path)
                img.width = w_px
                img.height = h_px
                ws_kapak.add_image(img, anchor_cell)

                cell = get_writable_cell(ws_kapak, caption_ref)
                cell.value = label_text
                cell.font = font_bold
                cell.alignment = align_center
            except Exception as e:
                sys.stderr.write(f"Warning: Failed to insert photo '{label_text}' at {anchor_cell}: {e}\n")


def clean_5070_text(wb):
    """
    Clears any cell value containing '5070' or 'elektronik imza kanunu' across all sheets.
    """
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    cval_u = cell.value.upper()
                    if "5070" in cval_u or "ELEKTRONİK İMZA KANUNU" in cval_u or "ELEKTRONIK IMZA KANUNU" in cval_u:
                        cell.value = None


def clean_left_personnel_block(wb, op_name: str, op_title: str, test_date: Optional[str] = None, report_date: Optional[str] = None, data_dict: Optional[Dict[str, Any]] = None):
    """
    Ensures KAPAK sol kutu has ONLY:
      - Test Tarihi : {test_date}
      - Rapor Tarihi : {report_date}
      - Unvan : {operator_title}
    (İsim Soyad row at A52 is removed/cleared as requested by user).

    Cleans and standardizes personnel blocks and footers across ALL sub-pages:
      - UNVAN: written verbatim from operator_title
      - TEST TARİHİ: test_date
      - ODA SİCİL NO / EKİPNET NO: written ONLY if valid sicil_no / ekipnet_no exists in data_dict, otherwise CLEARED to prevent showing operator_name/title.
    """
    if data_dict is None:
        data_dict = {}

    real_sicil = str(data_dict.get("sicil_no") or "").strip()
    real_ekipnet = str(data_dict.get("ekipnet_no") or "").strip()

    if real_sicil == op_name or real_sicil == op_title:
        real_sicil = ""
    if real_ekipnet == op_name or real_ekipnet == op_title:
        real_ekipnet = ""

    for sname in wb.sheetnames:
        ws = wb[sname]
        if sname == "KAPAK SAYFASI":
            # Clear A52:F52 (Remove operator_name / İsim Soyad from KAPAK left box)
            for c_let in ["A", "B", "C", "D", "E", "F"]:
                try:
                    get_writable_cell(ws, f"{c_let}52").value = None
                except Exception:
                    pass

            get_writable_cell(ws, "A53").value = "Test Tarihi"
            get_writable_cell(ws, "C53").value = ":"
            if test_date:
                s_val = date_to_excel_serial(test_date)
                cell_d53 = get_writable_cell(ws, "D53")
                cell_d53.value = s_val if s_val is not None else str(test_date)
                cell_d53.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "A54").value = "Rapor Tarihi"
            get_writable_cell(ws, "C54").value = ":"
            if report_date:
                s_val = date_to_excel_serial(report_date)
                cell_d54 = get_writable_cell(ws, "D54")
                cell_d54.value = s_val if s_val is not None else str(report_date)
                cell_d54.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "A55").value = "Unvan"
            get_writable_cell(ws, "C55").value = ":"
            get_writable_cell(ws, "D55").value = op_title

            # Clear A56:F57 residuals
            for r in [56, 57]:
                for c_let in ["A", "B", "C", "D", "E", "F"]:
                    try:
                        get_writable_cell(ws, f"{c_let}{r}").value = None
                    except Exception:
                        pass

        elif sname == "ANA SAYFA KESİCİ":
            get_writable_cell(ws, "B74").value = "Test Tarihi"
            get_writable_cell(ws, "E74").value = ":"
            if test_date:
                s_val = date_to_excel_serial(test_date)
                cell_f74 = get_writable_cell(ws, "F74")
                cell_f74.value = s_val if s_val is not None else str(test_date)
                cell_f74.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B75").value = "Rapor Tarihi"
            get_writable_cell(ws, "E75").value = ":"
            if report_date:
                s_val = date_to_excel_serial(report_date)
                cell_f75 = get_writable_cell(ws, "F75")
                cell_f75.value = s_val if s_val is not None else str(report_date)
                cell_f75.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B76").value = "Unvan"
            get_writable_cell(ws, "E76").value = ":"
            get_writable_cell(ws, "F76").value = op_title

            for r in [77, 78]:
                for c_let in ["A", "B", "C", "D", "E", "F"]:
                    try:
                        get_writable_cell(ws, f"{c_let}{r}").value = None
                    except Exception:
                        pass

        elif "ANA SAYFA" in sname and sname != "ANA SAYFA KESİCİ":
            get_writable_cell(ws, "B79").value = "Test Tarihi"
            get_writable_cell(ws, "E79").value = ":"
            if test_date:
                s_val = date_to_excel_serial(test_date)
                cell_f79 = get_writable_cell(ws, "F79")
                cell_f79.value = s_val if s_val is not None else str(test_date)
                cell_f79.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B80").value = "Rapor Tarihi"
            get_writable_cell(ws, "E80").value = ":"
            if report_date:
                s_val = date_to_excel_serial(report_date)
                cell_f80 = get_writable_cell(ws, "F80")
                cell_f80.value = s_val if s_val is not None else str(report_date)
                cell_f80.number_format = "dd.mm.yyyy"

            get_writable_cell(ws, "B81").value = "Unvan"
            get_writable_cell(ws, "E81").value = ":"
            get_writable_cell(ws, "F81").value = op_title

            # Clear rows 82, 83, 84 (A82:F84) - remove operator_name / İsim
            for r in [82, 83, 84]:
                for c_let in ["A", "B", "C", "D", "E", "F"]:
                    try:
                        get_writable_cell(ws, f"{c_let}{r}").value = None
                    except Exception:
                        pass

        # Sweep and clean sub-page personnel footers
        if sname != "KAPAK SAYFASI":
            for r in range(1, ws.max_row + 1):
                col_b_val = str(ws.cell(row=r, column=2).value or "").strip().upper()
                col_a_val = str(ws.cell(row=r, column=1).value or "").strip().upper()
                label_val = col_b_val if col_b_val else col_a_val
                if not label_val:
                    continue

                if label_val in ["UNVAN", "ÜNVAN"]:
                    target_cell = get_writable_cell(ws, f"F{r}")
                    target_cell.value = op_title
                elif label_val in ["TEST TARİHİ", "TEST TARIHI"]:
                    target_cell = get_writable_cell(ws, f"F{r}")
                    target_cell.value = "='KAPAK SAYFASI'!D53"
                    target_cell.number_format = "dd.mm.yyyy"
                elif any(kw in label_val for kw in ["ODA SİCİL NO", "ODA SICIL NO", "SICIL NO", "SİCİL NO", "EKİPNET NO", "EKIPNET NO"]):
                    # Clear entire row label & value cells for sicil_no / ekipnet_no
                    for c_let in ["A", "B", "C", "D", "E", "F"]:
                        try:
                            get_writable_cell(ws, f"{c_let}{r}").value = None
                        except Exception:
                            pass



def fix_subpage_dates(wb, test_date: Optional[str] = None, report_date: Optional[str] = None):
    """
    Fixes sub-page OTURUM TARİHİ cells across all worksheets (TOPRAKLAMALAR, HV PF, LV PF, DİĞER, AKIM TRAFOLARI, etc.)
    by pointing them to 'KAPAK SAYFASI'!D53 formula and applying number_format 'dd.mm.yyyy'.
    Ensures date cell never evaluates to 0 or '00.01.1900'.
    """
    subpage_date_cells = {
        "OG SARGI MEVCUT KADEME": "J9",
        "AG SARGI": "J9",
        "İZOLASYON ": "J9",
        "Ç.O 34500": "J9",
        "TOPRAKLAMALAR": "J7",
        "HV PF": "J9",
        "LV PF": "J9",
        "KESİCİ İZOLASYON": "J8",
        "KESİCİ KONTAK": "J8",
        "AÇMA-KAPAMA": "J7",
        "DİĞER": "J7",
        "AKIM TRAFOLARI": "J7",
    }
    for sname in wb.sheetnames:
        ws = wb[sname]
        cell_ref = subpage_date_cells.get(sname)
        if cell_ref:
            try:
                cell = get_writable_cell(ws, cell_ref)
                cell.value = "='KAPAK SAYFASI'!D53"
                cell.number_format = "dd.mm.yyyy"
            except Exception:
                pass
        else:
            for r in range(1, 15):
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if val is not None and ("OTURUM TARİHİ" in str(val).upper() or "OTURUM TARIHI" in str(val).upper()):
                        target_col = c + 8 if c == 2 else c + 2
                        target_cell = ws.cell(row=r, column=target_col)
                        if type(target_cell).__name__ != 'MergedCell':
                            target_cell.value = "='KAPAK SAYFASI'!D53"
                            target_cell.number_format = "dd.mm.yyyy"


def process_oil_petition(wb, data_dict):
    """
    Dynamically replaces sample terms (date, brand, power_kva, serial_no)
    in the body paragraph (B10) of 'HERMETİK YAĞ DİLEKÇESİ' sheet without altering template layout or legal text.
    Also clears residual D16/D18 values on petition sheet.
    """
    for sname in wb.sheetnames:
        if "YAĞ DİLEKÇESİ" in sname.upper() or "YAG DILEKCESI" in sname.upper():
            ws = wb[sname]

            # Clear D16 and D18 to eliminate '30,0' artifact
            for cref in ["D16", "D18"]:
                try:
                    get_writable_cell(ws, cref).value = None
                except Exception:
                    pass

            # Update customer/location headers
            try:
                c_name = data_dict.get("customer_name")
                if c_name:
                    get_writable_cell(ws, "J2").value = str(c_name).strip()
                else:
                    get_writable_cell(ws, "J2").value = "='KAPAK SAYFASI'!D9"

                addr = data_dict.get("address") or data_dict.get("location")
                if addr:
                    get_writable_cell(ws, "J3").value = str(addr).strip()
                else:
                    get_writable_cell(ws, "J3").value = "='KAPAK SAYFASI'!D11"
            except Exception:
                pass

            # Extract date
            raw_date = data_dict.get("test_date") or data_dict.get("report_date")
            formatted_date = None
            if raw_date:
                s_d = str(raw_date).strip()
                if "." in s_d:
                    formatted_date = s_d
                elif "-" in s_d:
                    parts = s_d.split("-")
                    if len(parts) == 3:
                        formatted_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
            if not formatted_date:
                formatted_date = datetime.now().strftime("%d.%m.%Y")

            brand_val = str(data_dict.get("brand") or "").strip()
            if not brand_val:
                brand_val = "ASTOR"

            power_val = data_dict.get("power_kva")
            if power_val:
                str_p = str(power_val).strip()
                if not str_p.lower().endswith("kva"):
                    str_p += "kVA"
            else:
                str_p = "1000kVA"

            serial_val = str(data_dict.get("serial_no") or "").strip()
            if not serial_val:
                serial_val = "27-07840"

            canonical_text = (
                "\n              İşletmenizde bulunan trafoların testini {date} tarihinde yapmış bulunmaktayız.\n"
                "Test kapsamında istenilen trafo yağ analizi işleminin yapılabilmesi için trafodan yağ numunesi alınması gerekmektedir. "
                "İşletmenizde bulunan {brand} marka, {power_kva}, '{serial_no}' seri numaralı trafonun hermetik tip olmasından dolayı yağ numunesi alınmamıştır.          \n"
                "                                                                                                                                                               \n"
                "              Hermetik tip trafolar atmosfere kapalıdır ve numune alma yerleri mühürlüdür. Mühürlerin açılıp numune alınması durumunda trafonun hermetik özelliği bozulacağından dolayı, fabrika standartlarına göre arıza durumu olmadığı sürece hiçbir şekilde hermetik trafolardan yağ numunesi alınamaz.\n\n"
                "              \n"
                "              Bilgilerinize sunarız.\n\n"
            )

            cell_b10 = get_writable_cell(ws, "B10")
            current_val = cell_b10.value

            if current_val and isinstance(current_val, str) and ("işletmenizde bulunan" in current_val.lower() or "trafo" in current_val.lower()):
                text = current_val
                text = re.sub(r"\b\d{2}\.\d{2}\.\d{4}\b", formatted_date, text)
                if brand_val != "ASTOR":
                    text = text.replace("ASTOR", brand_val)
                if str_p != "1000kVA":
                    text = text.replace("1000kVA", str_p).replace("1000 kVA", str_p)
                if serial_val != "27-07840":
                    text = text.replace("27-07840", serial_val)
                cell_b10.value = text
            else:
                text = canonical_text.format(
                    date=formatted_date,
                    brand=brand_val,
                    power_kva=str_p,
                    serial_no=serial_val
                )
                cell_b10.value = text

def apply_format_and_column_width_fixes(wb):
    """
    Ensures date cells, GΩ 20°C calculated cells, and V DC [V] cells have explicit number_format
    and sufficient column width to eliminate '#####' display issues.
    """
    # 1. KAPAK SAYFASI & ANA SAYFA (Col D for Dates)
    for sname in ["KAPAK SAYFASI", "ANA SAYFA"]:
        if sname in wb.sheetnames:
            ws = wb[sname]
            ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 14.0)
            for cref in ["D12", "D14", "D53", "D54"]:
                if cref in ws:
                    ws[cref].number_format = "dd.mm.yyyy"

    # 2. İZOLASYON sheet (Strict match: only 'İZOLASYON', NOT 'KESİCİ İZOLASYON')
    for sname in wb.sheetnames:
        norm_sname = sname.strip().upper()
        if norm_sname == "İZOLASYON":
            ws = wb[sname]
            for col_let in ["B", "D", "F", "H", "J", "L"]:
                ws.column_dimensions[col_let].width = max(ws.column_dimensions[col_let].width or 0, 13.0)
            for r in range(16, 32):
                for col_let in ["B", "D", "F", "H", "J", "L"]:
                    cell = ws[f"{col_let}{r}"]
                    if cell.value is not None:
                        if cell.number_format in ["General", "0.000000"]:
                            cell.number_format = "0.00"

    # 3. AG SARGI & OG SARGI MEVCUT KADEME (Col G for V DC [V])
    for sname in wb.sheetnames:
        if any(k in sname.upper() for k in ["AG SARGI", "OG SARGI"]):
            ws = wb[sname]
            ws.column_dimensions["G"].width = max(ws.column_dimensions["G"].width or 0, 15.0)
            for r in [24, 25, 26]:
                cell = ws[f"G{r}"]
                if cell.number_format in ["General", "0.000000"]:
                    cell.number_format = "0.00"


def resolve_default_template(mapped_type: str, repo_root: str) -> str:
    """
    Resolves template file path in priority order:
    1. TRAFO_TOOLS_DIR env -> ../backend/templates/hybrid/ or ../templates/hybrid/
    2. TRAFO_REPO_ROOT env -> backend/templates/hybrid/ or templates/hybrid/
    3. Relative repo_root (script parent) -> backend/templates/hybrid/
    4. CWD / Exe relative paths -> templates/hybrid/
    """
    type_filename_map = {
        "HERMETIK": ("hermetik_hybrid.xlsx", "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx", "hermetik.xlsx"),
        "GT": ("gt_hybrid.xlsx", "TR BAKIM RAPORU GT HİLMİ.xlsx", "gt.xlsx"),
        "KURU_TIP": ("kuru_tip_hybrid.xlsx", "KURU TİP HİLMİ.xlsx", "kuru_tip.xlsx"),
    }
    
    hybrid_fname, old_fname, mobile_fname = type_filename_map.get(mapped_type, type_filename_map["HERMETIK"])
    
    candidates = []

    env_tools_dir = os.environ.get("TRAFO_TOOLS_DIR")
    if env_tools_dir and os.path.isdir(env_tools_dir):
        candidates.append(os.path.abspath(os.path.join(env_tools_dir, "..", "backend", "templates", "hybrid", hybrid_fname)))
        candidates.append(os.path.abspath(os.path.join(env_tools_dir, "..", "templates", "hybrid", hybrid_fname)))
        candidates.append(os.path.abspath(os.path.join(env_tools_dir, "templates", "hybrid", hybrid_fname)))

    env_repo_root = os.environ.get("TRAFO_REPO_ROOT")
    if env_repo_root and os.path.isdir(env_repo_root):
        candidates.append(os.path.abspath(os.path.join(env_repo_root, "backend", "templates", "hybrid", hybrid_fname)))
        candidates.append(os.path.abspath(os.path.join(env_repo_root, "templates", "hybrid", hybrid_fname)))
        candidates.append(os.path.abspath(os.path.join(env_repo_root, "backend", "templates", old_fname)))

    candidates.extend([
        os.path.abspath(os.path.join(repo_root, "backend", "templates", "hybrid", hybrid_fname)),
        os.path.abspath(os.path.join(repo_root, "templates", "hybrid", hybrid_fname)),
        os.path.abspath(os.path.join(repo_root, "backend", "templates", old_fname)),
        os.path.abspath(os.path.join(repo_root, "mobile", "assets", "templates", mobile_fname)),
        os.path.abspath(os.path.join(os.getcwd(), "backend", "templates", "hybrid", hybrid_fname)),
        os.path.abspath(os.path.join(os.getcwd(), "templates", "hybrid", hybrid_fname)),
    ])

    for cand in candidates:
        if os.path.exists(cand):
            sys.stderr.write(f"DEBUG: Resolved template for {mapped_type}: {cand} (exists=True)\n")
            return cand

    sys.stderr.write(f"WARNING: No candidate template found for {mapped_type}, falling back to: {candidates[0]}\n")
    return candidates[0]


def main():
    parser = argparse.ArgumentParser(description="CLI Excel Generator for BTS Trafo Reports")
    parser.add_argument("--json", required=True, help="Path to form data JSON file")
    parser.add_argument("--template-type", required=True, help="Template type: hermetik | kuru_tip | gt")
    parser.add_argument("--template", required=False, help="Optional path to custom template .xlsx")
    parser.add_argument("--output", required=True, help="Target output .xlsx path")
    parser.add_argument("--signature", required=False, help="Optional signature PNG/JPG file path")
    parser.add_argument("--photo-before", required=False, help="Optional photo before path")
    parser.add_argument("--photo-after", required=False, help="Optional photo after path")
    parser.add_argument("--photo-label", required=False, help="Optional photo label path")
    parser.add_argument("--has-breaker", required=False, help="Explicit boolean flag for breaker sheets (true|false)")

    args = parser.parse_args()

    # 1. Read input JSON
    if not os.path.exists(args.json):
        sys.stderr.write(f"ERROR: Input JSON file not found: {args.json}\n")
        sys.exit(1)

    try:
        with open(args.json, "r", encoding="utf-8") as f:
            data_dict = json.load(f)
    except Exception as e:
        sys.stderr.write(f"ERROR: Failed to parse input JSON: {e}\n")
        sys.exit(1)

    # 2. Normalize template type
    raw_type = args.template_type.lower().strip()
    if "kuru" in raw_type:
        mapped_type = "KURU_TIP"
    elif "gt" in raw_type or "tank" in raw_type:
        mapped_type = "GT"
    else:
        mapped_type = "HERMETIK"

    # Set tank_mark checkmark based on type
    if mapped_type == "HERMETIK":
        data_dict["tank_mark_hermetik"] = "ü"
    elif mapped_type == "GT":
        data_dict["tank_mark_gt"] = "ü"
    elif mapped_type == "KURU_TIP":
        data_dict["tank_mark_kuru"] = "ü"

    # Address alias
    if "address" not in data_dict or not data_dict["address"]:
        data_dict["address"] = data_dict.get("location", "")

    # Normalize operator & title validation
    op_name = str(data_dict.get("operator_name") or data_dict.get("creator_display_name") or "").strip()
    if not op_name:
        sys.stderr.write("ERROR: Operatör adı (operator_name) eksik veya boş. Lütfen kullanıcı profil bilgilerinizi güncelleyin.\n")
        sys.exit(1)
    data_dict["operator_name"] = op_name

    op_title = str(data_dict.get("operator_title") or data_dict.get("title") or "").strip()
    if not op_title:
        sys.stderr.write("ERROR: Operatör unvanı (operator_title) eksik veya boş. Lütfen kullanıcı profil bilgilerinizi güncelleyin.\n")
        sys.exit(1)
    data_dict["operator_title"] = op_title

    if not data_dict.get("creator_display_name"):
        data_dict["creator_display_name"] = f"{op_name} ({op_title})"

    # Notes field normalization
    notes_text = data_dict.get("notes") or data_dict.get("notes_text") or ""
    if notes_text:
        notes_str = str(notes_text).strip()
        if not notes_str.upper().startswith("NOTLAR"):
            data_dict["notes"] = f"NOTLAR : {notes_str}"
        else:
            data_dict["notes"] = notes_str

    # Legacy nested aliases fallback
    wr = data_dict.get("winding_resistance") if isinstance(data_dict.get("winding_resistance"), dict) else {}
    data_dict.setdefault("og_rab", wr.get("r_phase"))
    data_dict.setdefault("og_rbc", wr.get("s_phase"))
    data_dict.setdefault("og_rca", wr.get("t_phase"))

    gr = data_dict.get("grounding") if isinstance(data_dict.get("grounding"), dict) else {}
    data_dict.setdefault("ground_r_trafo_body", data_dict.get("ground_trafo_body") or gr.get("value"))

    # Insulation 30s/45s/60s fallback aliases
    data_dict.setdefault("iso_yg_tank_30s", data_dict.get("iso_og_gnd_30s") or data_dict.get("iso_og_gnd") or data_dict.get("iso_yg_tank"))
    data_dict.setdefault("iso_yg_tank_45s", data_dict.get("iso_og_gnd_45s"))
    data_dict.setdefault("iso_yg_tank_60s", data_dict.get("iso_og_gnd_60s"))

    data_dict.setdefault("iso_ag_tank_30s", data_dict.get("iso_ag_gnd_30s") or data_dict.get("iso_ag_gnd") or data_dict.get("iso_ag_tank"))
    data_dict.setdefault("iso_ag_tank_45s", data_dict.get("iso_ag_gnd_45s"))
    data_dict.setdefault("iso_ag_tank_60s", data_dict.get("iso_ag_gnd_60s"))

    data_dict.setdefault("iso_yg_ag_30s", data_dict.get("iso_og_ag_30s") or data_dict.get("iso_og_ag") or data_dict.get("iso_yg_ag"))
    data_dict.setdefault("iso_yg_ag_45s", data_dict.get("iso_og_ag_45s"))
    data_dict.setdefault("iso_yg_ag_60s", data_dict.get("iso_og_ag_60s"))

    # Grounding 5-field fallback aliases
    data_dict.setdefault("ground_isletme_notr", data_dict.get("ground_r_neutral") or data_dict.get("ground_neutral"))
    data_dict.setdefault("ground_koruma_trafo", data_dict.get("ground_r_trafo_body") or data_dict.get("ground_trafo_body") or data_dict.get("ground_r_tank"))
    data_dict.setdefault("ground_koruma_hucre", data_dict.get("ground_r_hucre"))
    data_dict.setdefault("ground_koruma_kapilar", data_dict.get("ground_r_kapilar") or data_dict.get("ground_r_og_lightning"))
    data_dict.setdefault("ground_koruma_ag_pano", data_dict.get("ground_r_ag_pano") or data_dict.get("ground_r_panel"))

    # 3. Resolve template path
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    if args.template and os.path.exists(args.template):
        template_path = os.path.abspath(args.template)
    else:
        template_path = resolve_default_template(mapped_type, repo_root)

    if not os.path.exists(template_path):
        sys.stderr.write(f"ERROR: Template file not found: {template_path}\n")
        sys.exit(1)

    # 4. Load workbook without modifying template directly (data_only=False)
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
    except Exception as e:
        sys.stderr.write(f"ERROR: Failed to load workbook: {e}\n")
        sys.exit(1)

    cell_map_sheets = TYPE_CELL_MAPPINGS.get(mapped_type, TYPE_CELL_MAPPINGS["HERMETIK"])

    # 4.5. Pre-clear hardcoded sample data cells across all sheets
    pre_clear_sample_cells(wb)

    # 5. Cell writing loop across sheets (Mapping-first clear-then-write)
    force_overwrite_keys = {
        "operator_name", "operator_title", "creator_display_name", "notes", "device_model", "device_serial",
        "ground_isletme_notr", "ground_koruma_trafo", "ground_koruma_hucre", "ground_koruma_kapilar", "ground_koruma_ag_pano"
    }

    for target_sheet_name, cell_map in cell_map_sheets.items():
        matched_sheet = None
        for sname in wb.sheetnames:
            if sname.strip() == target_sheet_name.strip():
                matched_sheet = sname
                break

        if matched_sheet:
            ws = wb[matched_sheet]
            if target_sheet_name.strip() in ["ANA SAYFA", "ANA SAYFA KESİCİ"]:
                process_checklist_pairs(ws, mapped_type, data_dict)

            for cell_ref, field_key in cell_map.items():
                if field_key.startswith("TODO_VERIFY"):
                    continue

                cell_obj = get_writable_cell(ws, cell_ref)

                # Formula Protection: if existing cell value starts with '=', only overwrite if explicitly mapped personnel/notes key
                if cell_obj.value is not None and str(cell_obj.value).strip().startswith("="):
                    if field_key not in force_overwrite_keys:
                        continue

                val = data_dict.get(field_key)
                if val is None or val == "" or (isinstance(val, list) and len(val) == 0):
                    cell_obj.value = None
                    continue

                # Checkmarks: checklist_* or tank_mark_*
                if field_key.startswith("checklist_") or field_key.startswith("tank_mark_"):
                    if val is True or str(val).strip().lower() in ["true", "ü", "1", "evet"]:
                        cell_obj.value = "ü"
                    else:
                        cell_obj.value = None
                    continue

                # Dates
                if field_key in ["report_date", "test_date"]:
                    serial_val = date_to_excel_serial(val)
                    if serial_val is not None:
                        cell_obj.value = serial_val
                        cell_obj.number_format = "dd.mm.yyyy"
                    else:
                        cell_obj.value = None
                    continue

                # Numeric or string values
                if isinstance(val, bool):
                    cell_obj.value = "ü" if val else None
                elif isinstance(val, (int, float)):
                    cell_obj.value = val
                else:
                    str_val = str(val).strip()
                    if not str_val:
                        cell_obj.value = None
                    else:
                        clean_num = str_val.replace(',', '.')
                        try:
                            if '.' in clean_num:
                                cell_obj.value = float(clean_num)
                            else:
                                cell_obj.value = int(clean_num)
                        except Exception:
                            cell_obj.value = str_val

    # 6. Generalized Signature Image Cleaning across all sheets
    sig_path = args.signature or data_dict.get("signature_path") or data_dict.get("signature")
    sheet_signature_map = SHEET_SIGNATURE_ANCHORS.get(mapped_type, SHEET_SIGNATURE_ANCHORS["HERMETIK"])

    for sname in wb.sheetnames:
        ws = wb[sname]
        target_anchor = None
        for k_sheet, a_cell in sheet_signature_map.items():
            if k_sheet.strip() == sname.strip():
                target_anchor = a_cell
                break
        process_sheet_signature(ws, target_anchor, sig_path)

    # 7. Photos on KAPAK SAYFASI (Clean old background photos, insert user photos with Turkish captions)
    if "KAPAK SAYFASI" in wb.sheetnames:
        process_kapak_photos(wb["KAPAK SAYFASI"], data_dict, args)

    # 7.6. Clean left personnel block (ensure explicit labels and dates)
    test_date_val = data_dict.get("test_date")
    report_date_val = data_dict.get("report_date")
    clean_left_personnel_block(wb, op_name, op_title, test_date=test_date_val, report_date=report_date_val, data_dict=data_dict)
    fix_subpage_dates(wb, test_date=test_date_val, report_date=report_date_val)

    # 7.7. Process Oil Petition dynamic replacements (brand, date, power_kva, serial_no)
    process_oil_petition(wb, data_dict)

    # 8. Breaker Sheet Pruning (remove breaker sheets if has_breaker is False)
    BREAKER_SHEETS = ["ANA SAYFA KESİCİ", "KESİCİ İZOLASYON", "KESİCİ KONTAK", "AÇMA-KAPAMA"]
    has_breaker_arg = getattr(args, 'has_breaker', None)
    if has_breaker_arg is not None:
        is_breaker_enabled = str(has_breaker_arg).strip().lower() in ["true", "1", "yes", "evet"]
    else:
        has_breaker = data_dict.get("has_breaker")
        if has_breaker is None:
            has_breaker = data_dict.get("breaker_included")
        if has_breaker is None:
            has_breaker = bool(data_dict.get("breaker_brand") or data_dict.get("breaker_iso_r_gnd") or data_dict.get("breaker_contact_r"))
        is_breaker_enabled = (has_breaker is True or str(has_breaker).strip().lower() in ["true", "1", "yes", "evet"])

    data_dict["has_breaker"] = is_breaker_enabled
    data_dict["breaker_included"] = is_breaker_enabled

    if not is_breaker_enabled:
        for sname in list(wb.sheetnames):
            norm_sname = sname.strip().upper()
            for b_sheet in BREAKER_SHEETS:
                if norm_sname == b_sheet.strip().upper():
                    wb.remove(wb[sname])
                    break

    # 9. Fail-safe text sweeper to clear any lingering sample strings (ULUSOY, USFB-36R, Hilmi, 88258, K202439698)
    b_brand = str(data_dict.get("breaker_brand") or data_dict.get("brand") or "").strip()
    b_model = str(data_dict.get("breaker_model") or "").strip()

    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if type(cell).__name__ == 'MergedCell':
                    continue
                if cell.value is not None and isinstance(cell.value, str):
                    cval = cell.value.strip()
                    if cval.startswith('='):
                        continue
                    cval_u = cval.upper()

                    if "HİLMİ" in cval_u or "HILMI" in cval_u:
                        for old_term in ["Hilmi GÜL", "Hilmi GUL", "Hilmi"]:
                            if old_term.upper() in cval_u:
                                cval = cval.replace(old_term, op_name) if op_name else ""
                    if "ULUSOY" in cval_u:
                        cval = cval.replace("ULUSOY", b_brand) if b_brand else ""
                        cval = cval.replace("ULusoy", b_brand) if b_brand else ""
                    if "USFB-36R" in cval_u:
                        cval = cval.replace("USFB-36R", b_model) if b_model else ""
                    if "K202439698" in cval_u or "88258" in cval_u:
                        cval = ""

                    cell.value = cval if cval.strip() != "" else None

    # 10. Apply column width & number format fixes to prevent '#####' issues
    apply_format_and_column_width_fixes(wb)

    # 11. Save output file
    output_abs_path = os.path.abspath(args.output)
    output_dir = os.path.dirname(output_abs_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    try:
        wb.save(output_abs_path)
        fix_xlsx_rels(output_abs_path)
    except Exception as e:
        sys.stderr.write(f"ERROR: Failed to save output workbook: {e}\n")
        sys.exit(1)

    # 12. Success stdout signature
    print(f"OUTPUT_OK:{output_abs_path}")
    sys.exit(0)


def fix_xlsx_rels(xlsx_path: str):
    """
    Fixes relationship target paths in .xlsx zip archive by converting
    absolute targets starting with '/xl/' (e.g. '/xl/drawings/drawing1.xml', '/xl/media/image1.png')
    to standard relative targets without leading '/xl/'.
    """
    try:
        import io, zipfile
        import xml.etree.ElementTree as ET

        buffer = io.BytesIO()
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    content = zin.read(item.filename)
                    if item.filename.endswith('.rels'):
                        try:
                            root = ET.fromstring(content)
                            modified = False
                            parts = item.filename.split('/')
                            if len(parts) >= 2 and parts[-2] == '_rels':
                                base_dir = '/'.join(parts[:-2])
                            else:
                                base_dir = '/'.join(parts[:-1])

                            for child in root:
                                tgt = child.get('Target')
                                if tgt and tgt.startswith('/xl/'):
                                    full_target = tgt[1:]  # strip leading '/'
                                    rel_target = os.path.relpath(full_target, base_dir).replace('\\', '/')
                                    child.set('Target', rel_target)
                                    modified = True

                            if modified:
                                content = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                        except Exception:
                            pass
                    zout.writestr(item, content)

        with open(xlsx_path, 'wb') as f:
            f.write(buffer.getvalue())
    except Exception as e:
        sys.stderr.write(f"Warning: Failed to fix rels targets in '{xlsx_path}': {e}\n")


if __name__ == "__main__":
    main()


if __name__ == "__main__":
    main()
