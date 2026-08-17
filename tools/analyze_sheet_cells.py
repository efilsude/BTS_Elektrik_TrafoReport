import os
import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TEMPLATES_DIR = os.path.join(REPO_ROOT, "backend", "templates")
HERMETIK_PATH = os.path.join(TEMPLATES_DIR, "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx")

def main():
    wb = openpyxl.load_workbook(HERMETIK_PATH, data_only=False)
    for sname in wb.sheetnames:
        ws = wb[sname]
        data_cells = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    val_str = str(cell.value).strip()
                    if not val_str.startswith('='):
                        col_let = openpyxl.utils.get_column_letter(c)
                        safe_val = val_str.encode('ascii', 'backslashreplace').decode('ascii')
                        data_cells.append(f"{col_let}{r}:{safe_val}")
        print(f"=== {sname} ({len(data_cells)} cells) ===")
        print(" | ".join(data_cells))
        print("\n" + "="*80 + "\n")

if __name__ == "__main__":
    main()
