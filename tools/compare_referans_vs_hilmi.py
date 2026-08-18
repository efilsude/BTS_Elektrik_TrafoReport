import os
import openpyxl
import zipfile
import xml.etree.ElementTree as ET

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TEMPLATES_DIR = os.path.join(REPO_ROOT, "backend", "templates")
REF_PATH = os.path.join(TEMPLATES_DIR, "reference", "yeni_tasarim_referans.xlsx")
HILMI_PATH = os.path.join(TEMPLATES_DIR, "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx")
HYBRID_PATH = os.path.join(TEMPLATES_DIR, "hybrid", "hermetik_hybrid.xlsx")

def print_sheet_details(wb_name, wb):
    print(f"\n==================================================")
    print(f"WORKBOOK: {wb_name}")
    print(f"==================================================")
    print(f"Sheet names: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        sf = getattr(ws, 'sheet_format', None)
        def_rh = getattr(sf, 'defaultRowHeight', None)
        def_cw = getattr(sf, 'defaultColWidth', None)
        max_r = ws.max_row
        max_c = ws.max_column
        
        # Check column widths A..R
        col_w = {}
        for col_idx in range(1, min(20, max_c + 1)):
            c_let = openpyxl.utils.get_column_letter(col_idx)
            cdim = ws.column_dimensions.get(c_let)
            w = cdim.width if cdim else None
            col_w[c_let] = w
            
        # Check sample row heights
        rh_sample = []
        for r in range(1, min(25, max_r + 1)):
            rdim = ws.row_dimensions.get(r)
            rh_sample.append((r, rdim.height if rdim else None))
            
        print(f"\n--- Sheet: {sname} ---")
        print(f"Dimensions: max_row={max_r}, max_col={max_c}")
        print(f"sheet_format: defaultRowHeight={def_rh}, defaultColWidth={def_cw}")
        print(f"Col Widths (A..N): {col_w}")
        print(f"Row Heights (1..15): {rh_sample[:15]}")

def check_xml_views(xlsx_path):
    print(f"\n--- RAW XML VIEWS FOR {os.path.basename(xlsx_path)} ---")
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        wb_xml = z.read('xl/workbook.xml')
        wb_tree = ET.fromstring(wb_xml)
        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
              'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        
        rel_map = {}
        wb_rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_tree = ET.fromstring(wb_rels_xml)
        for child in rels_tree:
            rid = child.get('Id')
            target = child.get('Target')
            rel_map[rid] = target

        for s in wb_tree.findall('.//ns:sheet', ns):
            s_name = s.get('name')
            s_rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            target_path = rel_map.get(s_rid, '')
            if target_path.startswith('/xl/'):
                target_path = target_path[4:]
            elif not target_path.startswith('xl/'):
                target_path = 'xl/' + target_path
            
            xml_content = z.read(target_path)
            tree = ET.fromstring(xml_content)
            sv = tree.find('.//ns:sheetView', ns)
            view_mode = sv.get('view') if sv is not None else None
            zoom = sv.get('zoomScale') if sv is not None else None
            print(f"Sheet '{s_name}': xml_path={target_path}, view_mode={view_mode}, zoomScale={zoom}")

def main():
    wb_ref = openpyxl.load_workbook(REF_PATH, data_only=False)
    wb_hilmi = openpyxl.load_workbook(HILMI_PATH, data_only=False)
    
    print_sheet_details("yeni_tasarim_referans.xlsx", wb_ref)
    print_sheet_details("HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx", wb_hilmi)
    
    check_xml_views(REF_PATH)
    check_xml_views(HILMI_PATH)
    check_xml_views(HYBRID_PATH)

if __name__ == "__main__":
    main()
