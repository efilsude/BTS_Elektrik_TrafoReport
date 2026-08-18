import os
import sys
import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
BACKEND_DIR = os.path.join(REPO_ROOT, "backend")
sys.path.insert(0, BACKEND_DIR)

from app.services.excel_engine import generate_excel_report

class MockReport:
    def __init__(self, report_type, data_dict):
        self.report_type = report_type
        self.transformer_type = report_type
        self.customer_name = data_dict.get("customer_name", "CHECKLIST TEST A.S.")
        self.trafo_label = data_dict.get("trafo_label", "TR-CHK-01")
        self.data_json = data_dict

def test_checklist_system():
    types_to_test = ["HERMETIK", "KURU_TIP", "GT"]
    all_passed = True

    # Test inputs
    test_cases = {
        "HERMETIK": {
            "dc_redresor_voltage": None,
            "checklist_1": True,
            "checklist_2": False,
            "checklist_3": True,
            "checklist_4": False,
            "checklist_5": True,
            "checklist_6": False,
            "checklist_7": True,
            "checklist_8": False,
            "checklist_9": True,
            "checklist_10": False,
            "checklist_11": True,
            "checklist_12": False,
            "checklist_13": True,
            "checklist_14": False,
            "checklist_15": True,
            "checklist_16": False,
        },
        "KURU_TIP": {
            "dc_redresor_voltage": "24 VDC",
            "checklist_1": False,
            "checklist_2": True,
            "checklist_3": False,
            "checklist_4": True,
            "checklist_5": False,
            "checklist_6": True,
            "checklist_7": False,
            "checklist_8": True,
            "checklist_9": False,
            "checklist_10": True,
            "checklist_11": False,
            "checklist_12": True,
        },
        "GT": {
            "dc_redresor_voltage": "110 VDC",
            "checklist_1": True,
            "checklist_2": True,
            "checklist_3": True,
            "checklist_4": True,
            "checklist_5": False,
            "checklist_6": False,
            "checklist_7": False,
            "checklist_8": False,
            "checklist_9": True,
            "checklist_10": True,
            "checklist_11": True,
            "checklist_12": True,
            "checklist_13": False,
            "checklist_14": False,
            "checklist_15": False,
            "checklist_16": False,
        }
    }

    # Expected checklist pairs coordinate maps
    hermetik_gt_pairs = {
        "checklist_1": ("I27", "J27"),
        "checklist_2": ("I29", "J29"),
        "checklist_3": ("I31", "J31"),
        "checklist_4": ("I33", "J33"),
        "checklist_5": ("I35", "J35"),
        "checklist_6": ("I37", "J37"),
        "checklist_7": ("I39", "J39"),
        "checklist_8": ("I41", "J41"),
        "checklist_9": ("R27", "S27"),
        "checklist_10": ("R29", "S29"),
        "checklist_11": ("R31", "S31"),
        "checklist_12": ("R33", "S33"),
        "checklist_13": ("R35", "S35"),
        "checklist_14": ("R37", "S37"),
        "checklist_15": ("R39", "S39"),
        "checklist_16": ("R41", "S41"),
    }

    kuru_pairs = {
        "checklist_1": ("I27", "J27"),
        "checklist_2": ("I29", "J29"),
        "checklist_3": ("I31", "J31"),
        "checklist_4": ("I33", "J33"),
        "checklist_5": ("I35", "J35"),
        "checklist_6": ("I37", "J37"),
        "checklist_7": ("R27", "S27"),
        "checklist_8": ("R29", "S29"),
        "checklist_9": ("R31", "S31"),
        "checklist_10": ("R33", "S33"),
        "checklist_11": ("R35", "S35"),
        "checklist_12": ("R37", "S37"),
    }

    for rtype in types_to_test:
        print(f"\n==================================================")
        print(f"TESTING CHECKLIST FOR {rtype}")
        print(f"==================================================")
        data = test_cases[rtype]
        out_file = os.path.join(REPO_ROOT, "tools", f"test_checklist_out_{rtype.lower()}.xlsx")
        report = MockReport(rtype, data)
        generate_excel_report(report, photos=[], output_path=out_file)
        
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws = wb['ANA SAYFA']

        # 1. Verify DC Redresor Voltage at G31
        g31_val = ws['G31'].value
        expected_g31 = data['dc_redresor_voltage']
        print(f"[{rtype}] Cell G31 (DC Redresor Voltage): {g31_val!r} (Expected: {expected_g31!r})")
        if g31_val != expected_g31:
            print(f"  FAIL: G31 mismatch in {rtype}! Got {g31_val!r}, expected {expected_g31!r}")
            all_passed = False

        # 2. Verify Checklist Pairs
        pairs = kuru_pairs if rtype == "KURU_TIP" else hermetik_gt_pairs
        for item_key, (evet_ref, hayir_ref) in pairs.items():
            evet_val = ws[evet_ref].value
            hayir_val = ws[hayir_ref].value
            expected_choice = data[item_key]
            
            if expected_choice is True:
                expected_evet, expected_hayir = "ü", None
            else:
                expected_evet, expected_hayir = None, "ü"
                
            is_ok = (evet_val == expected_evet and hayir_val == expected_hayir)
            print(f"[{rtype}] {item_key} ({evet_ref}/{hayir_ref}): Evet={evet_val!r}, Hayir={hayir_val!r} -> {'OK' if is_ok else 'FAIL (Expected Evet=' + repr(expected_evet) + ', Hayir=' + repr(expected_hayir) + ')'}")
            if not is_ok:
                all_passed = False

        # 3. For Kuru Tip, verify rows 39 & 41 have NO checklist items or checkmarks
        if rtype == "KURU_TIP":
            b39, k39 = ws['B39'].value, ws['K39'].value
            b41, k41 = ws['B41'].value, ws['K41'].value
            i39, j39 = ws['I39'].value, ws['J39'].value
            i41, j41 = ws['I41'].value, ws['J41'].value
            r39, s39 = ws['R39'].value, ws['S39'].value
            r41, s41 = ws['R41'].value, ws['S41'].value
            no_oil_pressure_gaskets = (
                b39 is None and k39 is None and b41 is None and k41 is None and
                i39 is None and j39 is None and i41 is None and j41 is None and
                r39 is None and s39 is None and r41 is None and s41 is None
            )
            print(f"[{rtype}] Rows 39 & 41 empty (no oil/pressure/gasket items): {'OK' if no_oil_pressure_gaskets else 'FAIL'}")
            if not no_oil_pressure_gaskets:
                all_passed = False

    print("\n==================================================")
    print(f"ALL CHECKLIST SYSTEM TESTS PASSED: {all_passed}")
    print(f"==================================================")
    return all_passed

if __name__ == "__main__":
    if not test_checklist_system():
        sys.exit(1)
