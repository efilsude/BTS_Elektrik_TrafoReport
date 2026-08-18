import os
import sys
import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
BACKEND_DIR = os.path.join(REPO_ROOT, "backend")
sys.path.insert(0, BACKEND_DIR)

from app.services.excel_engine import generate_excel_report

class MockReport:
    def __init__(self, report_type, data_dict):
        self.report_type = report_type
        self.transformer_type = report_type
        self.customer_name = data_dict.get("customer_name", "TOPRAKLAMA TEST A.S.")
        self.trafo_label = data_dict.get("trafo_label", "TR-4OHM-01")
        self.data_json = data_dict

def test_grounding_limit_update():
    files_to_verify = [
        'backend/templates/HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx',
        'backend/templates/KURU TİP HİLMİ.xlsx',
        'backend/templates/TR BAKIM RAPORU GT HİLMİ.xlsx',
        'backend/templates/hybrid/gt_hybrid.xlsx',
        'backend/templates/hybrid/hermetik_hybrid.xlsx',
        'backend/templates/hybrid/kuru_tip_hybrid.xlsx',
        'mobile/assets/templates/gt.xlsx',
        'mobile/assets/templates/gt_hybrid.xlsx',
        'mobile/assets/templates/hermetik.xlsx',
        'mobile/assets/templates/hermetik_hybrid.xlsx',
        'mobile/assets/templates/kuru_tip.xlsx',
        'mobile/assets/templates/kuru_tip_hybrid.xlsx',
    ]

    all_passed = True
    cells = ['D16', 'F26', 'F27', 'F28', 'F29']

    print("=== Step 1: Programmatic Verification of 12 Template Files ===")
    for rel_path in files_to_verify:
        fpath = os.path.join(REPO_ROOT, rel_path)
        if not os.path.exists(fpath):
            print(f"FAIL: File not found {rel_path}")
            all_passed = False
            continue
        wb = openpyxl.load_workbook(fpath)
        ws = wb['TOPRAKLAMALAR']
        vals = {c: ws[c].value for c in cells}
        is_ok = all(v == 4 for v in vals.values())
        print(f"File '{rel_path}': {vals} -> {'OK' if is_ok else 'FAIL'}")
        if not is_ok:
            all_passed = False

    print("\n=== Step 2: Generating Excel Reports with Grounding Resistance = 3 Ohm ===")
    types = ["HERMETIK", "KURU_TIP", "GT"]
    data = {
        "customer_name": "TOPRAKLAMA TEST A.S.",
        "trafo_label": "TR-4OHM-01",
        "ground_r_trafo_body": "3.0",
        "ground_r_neutral": "3.0",
        "ground_r_tank": "3.0",
        "ground_r_og_lightning": "3.0",
        "ground_r_panel": "3.0",
        "ground_r_fence": "3.0",
    }

    for rtype in types:
        out_file = os.path.join(REPO_ROOT, "tools", f"test_grounding_4ohm_{rtype.lower()}.xlsx")
        report = MockReport(rtype, data)
        generate_excel_report(report, photos=[], output_path=out_file)
        wb = openpyxl.load_workbook(out_file, data_only=False)
        ws = wb['TOPRAKLAMALAR']
        gen_vals = {c: ws[c].value for c in cells}
        print(f"Generated {rtype} [{out_file}]: TOPRAKLAMALAR limits = {gen_vals}")
        if not all(v == 4 for v in gen_vals.values()):
            print(f"FAIL: Generated report {rtype} limits are not 4!")
            all_passed = False

    print("\n==================================================")
    print(f"GROUNDING LIMIT UPDATE VERIFICATION PASSED: {all_passed}")
    print(f"==================================================")
    return all_passed

if __name__ == "__main__":
    if not test_grounding_limit_update():
        sys.exit(1)
