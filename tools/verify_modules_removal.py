import os
import sys
import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
BACKEND_DIR = os.path.join(REPO_ROOT, "backend")
sys.path.insert(0, BACKEND_DIR)

from app.services.excel_engine import generate_excel_report

class MockReport:
    def __init__(self, report_type, data_dict):
        self.report_type = report_type
        self.transformer_type = report_type
        self.customer_name = data_dict.get("customer_name", "MODULSIZ TEST A.S.")
        self.trafo_label = data_dict.get("trafo_label", "TR-TEST-01")
        self.data_json = data_dict

def test_modules_removal_in_excel():
    types_to_test = ["HERMETIK", "KURU_TIP", "GT"]
    
    # Input data dictionary WITHOUT module fields (pf_hv_humidity, pf_lv_humidity, ct_ratio)
    data = {
        "customer_name": "MODULSIZ TEST A.S.",
        "trafo_label": "TR-TEST-01",
        "address": "Sanayi Mah. 100. Yil Cad.",
        "test_date": "18.08.2026",
        "report_date": "18.08.2026",
        "operator_name": "Ahmet Yilmaz",
        "operator_title": "Test Muhendisi",
        "brand": "ASTOR",
        "power_kva": "1000kVA",
        "serial_no": "AST-2026-10",
        "has_breaker": True,
        "breaker_brand": "SIEMENS",
    }

    all_passed = True

    for rtype in types_to_test:
        print(f"\n--- Verifying Excel output for {rtype} without Modules ---")
        output_file = os.path.join(REPO_ROOT, "tools", f"test_no_modules_{rtype.lower()}.xlsx")
        report = MockReport(rtype, data)
        
        generate_excel_report(report, photos=[], output_path=output_file)
        print(f"Generated: {output_file}")
        
        wb = openpyxl.load_workbook(output_file, data_only=False)
        sheet_names = wb.sheetnames
        print(f"Sheet count: {len(sheet_names)}")
        
        # Check expected module sheets existence
        if rtype in ["HERMETIK", "KURU_TIP"]:
            target_sheets = ["HV PF", "LV PF", "AKIM TRAFOLARI"]
        else: # GT template does not include PF or AKIM TRAFOLARI sheets
            target_sheets = []
            
        for sname in target_sheets:
            if sname not in sheet_names:
                print(f"FAIL: Sheet '{sname}' missing from {rtype} output!")
                all_passed = False
            else:
                ws = wb[sname]
                print(f"Sheet '{sname}' exists.")
                
                # Verify specific data cells are empty / cleared
                if sname in ["HV PF", "LV PF"]:
                    cell_val = ws["P17"].value
                    print(f"  [{sname}] P17 value: {cell_val!r}")
                    if cell_val is not None:
                        print(f"  FAIL: Expected P17 to be None in '{sname}', got {cell_val!r}")
                        all_passed = False
                elif sname == "AKIM TRAFOLARI":
                    cell_val = ws["D16"].value
                    print(f"  [{sname}] D16 value: {cell_val!r}")
                    if cell_val is not None:
                        print(f"  FAIL: Expected D16 to be None in '{sname}', got {cell_val!r}")
                        all_passed = False

    print(f"\n==================================================")
    print(f"MODULES REMOVAL EXCEL VERIFICATION PASSED: {all_passed}")
    print(f"==================================================")
    return all_passed

if __name__ == "__main__":
    if not test_modules_removal_in_excel():
        sys.exit(1)
