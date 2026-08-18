import os
import openpyxl
import zipfile
import xml.etree.ElementTree as ET

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TEMPLATES_DIR = os.path.join(REPO_ROOT, "backend", "templates")
REF_PATH = os.path.join(TEMPLATES_DIR, "reference", "yeni_tasarim_referans.xlsx")
HILMI_PATH = os.path.join(TEMPLATES_DIR, "HERMETİK TRAFO BAKIM RAPORU HİLMİ.xlsx")
HYBRID_PATH = os.path.join(TEMPLATES_DIR, "hybrid", "hermetik_hybrid.xlsx")
OUTPUT_PATH = os.path.join(REPO_ROOT, "tools", "t1_minimal_output.xlsx")

def inspect_sheet_full(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    matched = None
    for s in wb.sheetnames:
        if s.strip() == sheet_name.strip():
            matched = s
            break
    if not matched:
        return f"Sheet {sheet_name} not found in {os.path.basename(xlsx_path)}"
    
    ws = wb[matched]
    
    # 1. Total Column Width (A to R)
    col_widths = {}
    total_w = 0.0
    for c in range(1, 20):
        c_let = openpyxl.utils.get_column_letter(c)
        cdim = ws.column_dimensions.get(c_let)
        w = cdim.width if cdim and cdim.width is not None else 8.43
        col_widths[c_let] = w
        total_w += w
        
    # 2. Total Row Height (1 to max_row)
    total_h = 0.0
    explicit_h_count = 0
    missing_h_count = 0
    rh_list = []
    for r in range(1, ws.max_row + 1):
        rdim = ws.row_dimensions.get(r)
        h = rdim.height if rdim and rdim.height is not None else 15.0
        if rdim and rdim.height is not None:
            explicit_h_count += 1
        else:
            missing_h_count += 1
        total_h += h
        if r <= 30:
            rh_list.append(h)
            
    # 3. Fonts & Styles sample
    fonts = set()
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), min_col=1, max_col=15):
        for cell in row:
            if cell.font:
                fonts.add((cell.font.name, cell.font.size, cell.font.bold))
                
    # 4. XML specifics
    xml_info = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        wb_xml = z.read('xl/workbook.xml')
        wb_tree = ET.fromstring(wb_xml)
        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
              'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        
        rel_map = {}
        wb_rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_tree = ET.fromstring(wb_rels_xml)
        for child in rels_tree:
            rel_map[child.get('Id')] = child.get('Target')

        for s in wb_tree.findall('.//ns:sheet', ns):
            if s.get('name').strip() == matched.strip():
                target_path = rel_map.get(s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'), '')
                if target_path.startswith('/xl/'):
                    target_path = target_path[4:]
                elif not target_path.startswith('xl/'):
                    target_path = 'xl/' + target_path
                
                sheet_xml = z.read(target_path)
                tree = ET.fromstring(sheet_xml)
                
                # Check pageSetup, sheetView, sheetFormatPr, printOptions, pageMargins
                ps = tree.find('.//ns:pageSetup', ns)
                pspr = tree.find('.//ns:pageSetUpPr', ns)
                sv = tree.find('.//ns:sheetView', ns)
                sf = tree.find('.//ns:sheetFormatPr', ns)
                
                xml_info['pageSetup'] = {k: ps.get(k) for k in ps.attrib} if ps is not None else None
                xml_info['pageSetUpPr'] = {k: pspr.get(k) for k in pspr.attrib} if pspr is not None else None
                xml_info['sheetView'] = {k: sv.get(k) for k in sv.attrib} if sv is not None else None
                xml_info['sheetFormatPr'] = {k: sf.get(k) for k in sf.attrib} if sf is not None else None

    return {
        'sheet': matched,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'total_width_A_R': round(total_w, 2),
        'total_height': round(total_h, 2),
        'explicit_h_count': explicit_h_count,
        'missing_h_count': missing_h_count,
        'col_widths_sample': dict(list(col_widths.items())[:10]),
        'first_15_row_heights': rh_list[:15],
        'fonts': list(fonts)[:5],
        'xml_info': xml_info
    }

def main():
    sheets_to_compare = [
        ("yeni_tasarim_referans.xlsx", REF_PATH, "KAPAK SAYFASI"),
        ("yeni_tasarim_referans.xlsx", REF_PATH, "ANA SAYFA"),
        ("HİLMİ.xlsx", HILMI_PATH, "ANA SAYFA"),
        ("HİLMİ.xlsx", HILMI_PATH, "ANA SAYFA KESİCİ"),
        ("hermetik_hybrid.xlsx", HYBRID_PATH, "ANA SAYFA KESİCİ"),
        ("t1_minimal_output.xlsx", OUTPUT_PATH, "ANA SAYFA KESİCİ"),
        ("HİLMİ.xlsx", HILMI_PATH, "KESİCİ İZOLASYON"),
        ("hermetik_hybrid.xlsx", HYBRID_PATH, "KESİCİ İZOLASYON"),
        ("t1_minimal_output.xlsx", OUTPUT_PATH, "KESİCİ İZOLASYON"),
    ]
    
    for tag, path, sname in sheets_to_compare:
        print(f"\n==================================================")
        print(f"{tag} -> {sname}")
        print(f"==================================================")
        res = inspect_sheet_full(path, sname)
        if isinstance(res, dict):
            for k, v in res.items():
                print(f"  {k}: {v}")
        else:
            print(res)

if __name__ == "__main__":
    main()
